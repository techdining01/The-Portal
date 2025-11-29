from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.db.models import Sum, Avg, Count, Min, Max, Q, Prefetch
from django.template.loader import render_to_string
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST, require_http_methods
import json, os, io, datetime
from io import BytesIO
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from django.db import transaction
from django.contrib.auth import get_user_model
from .models import Quiz, Question, Choice, StudentQuizAttempt, ActionLog, Answer, Class, Subject,RetakeRequest
from users.models import Notification
from .utils import log_action
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import Avg
from django.db.models import Sum
from django.utils import timezone
from reportlab.platypus import PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import matplotlib
import matplotlib.pyplot as plt

matplotlib.use('Agg')  

User = get_user_model()


# --------------------------------------------------------------#

def is_teacher(user):
    return user.is_authenticated and user.role == 'teacher'

def is_admin(user):
    return user.is_authenticated and user.role in ('admin', 'superadmin')

def is_teacher_or_admin(user):
    return user.is_authenticated and user.role in ('teacher', 'admin', 'superadmin')

def is_admin_or_superadmin(user):
    return user.is_authenticated and user.role in ["admin", "superadmin"]


def home(request):
    return render(request, "exams/home.html")


@login_required
@user_passes_test(is_admin_or_superadmin)
def superadmin_dashboard(request):
    if request.user.role !=  "superadmin":
        return HttpResponseForbidden("Unauthorized")
    # counts
    approved = User.objects.filter(role__in=['admin', 'teacher', 'student'], approved=True).count()
    pended = User.objects.filter(approved=False).count()
    teachers = User.objects.filter(role='teacher', approved=True).count()
    students = User.objects.filter(role='student', approved=True).count()

    # leaderboard: top students by average total_score (only graded attempts)
    leaderboard = (
        StudentQuizAttempt.objects.filter(is_submitted=True)
        .values("student__username")
        .annotate(avg_score=Avg("score"))
        .order_by("-avg_score")[:10]  # top 10
    )

    # recent action logs
    actions = ActionLog.objects.order_by('-timestamp')[:20]

    notifications = Notification.objects.filter(role='superadmin').order_by('-created_at')[:10]
    

    context = {
        "approved": approved,
        "pended": pended,
        "teachers": teachers,
        "students": students,
        "leaderboard": leaderboard,
        "actions": actions,
        "notifications": notifications,
        "total_admins": User.objects.filter(role="admin", approved=True).count(),
        "total_quizzes": Quiz.objects.count(),
        "total_classes": Class.objects.count(),
        'total_subjects': Subject.objects.count(),
    }
    return render(request, "exams/superadmin_dashboard.html", context)


# helper: admin check
def is_admin(user):
    return user.is_authenticated and user.role in ("admin", "superadmin")

# Admin dashboard page (HTML shell)
@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    admin = request.user
    User = get_user_model()
    # Deducting number of admin and teacher registered for every classes from total student per class
    classes = Class.objects.prefetch_related(
    Prefetch('student_class', queryset=User.objects.filter(role='student')))

    # students that have at least one submitted attempt in this teacher's class
    students_with_attempts_qs = User.objects.filter(
        studentquizattempt__is_submitted=True,
        studentquizattempt__student__student_class=admin.student_class,
        role="student"
    ).distinct().order_by('last_name', 'first_name', 'username')

    # paginate students list for download/report links on dashboard
    students_page_num = int(request.GET.get("students_with_att_page", 1) or 1)
    students_per_page = 7  
    students_with_att_page = Paginator(students_with_attempts_qs, students_per_page).get_page(students_page_num)

    # pending subjective attempts for grading (existing behavior)
    pending_attempts = (
        StudentQuizAttempt.objects
        .filter(
            quiz__created_by=admin,
            is_submitted=True,
            answers__is_pending=True,
            answers__question__question_type='subjective'
        )
        .select_related('quiz', 'student')
        .distinct()
        .order_by('-submitted_at')
    )


      # Show pending attempts initially (attempts that have at least one pending subjective answer)
    setter = User.objects.get(id=request.user.id, role__in=["admin", "superadmin"])
    pending_attempts = StudentQuizAttempt.objects.filter(quiz__created_by=setter, is_submitted=True, answers__is_pending=True).select_related('quiz', 'student').distinct().order_by('-submitted_at')
    grade_page_num = request.GET.get("grade_page", 1)

    grade_page = Paginator(pending_attempts, 5).get_page(grade_page_num)
   
    grading = [
    {
        "id": a.id,
        "student": a.student.username,
        "full_name": a.student.get_full_name(),
        "quiz": a.quiz.title,
        "quiz_id": a.quiz.id,
        "score": float(a.score or 0.0),
        "graded": bool(a.graded),
        "submitted_at": datetime.date(a.submitted_at) if getattr(a, "submitted_at", None) else None,
    }
    for a in grade_page
    ]
    
    context = {
        "attempts": pending_attempts,
        # keep both full queryset and paginated page (template can use page object)
        "students_with_attempts": students_with_attempts_qs,
        "students_with_att_page": students_with_att_page,
        "students_with_att_num": students_page_num,
        "students_with_att_num_pages": students_with_att_page.paginator.num_pages,
        'classes': classes, 
        "grading": grading,
    }
      
    return render(request, "exams/admin_dashboard.html", context )


# Admin dashboard data endpoint (GET -> fetch data; POST -> perform actions such as approve/reject/broadcast/download)
@login_required
def admin_dashboard_data(request):
    if request.user.role not in ("admin", "superadmin"):
        return JsonResponse({"error": "forbidden"}, status=403)

    # Handle POST actions: approve/reject/pending user, broadcast, download, etc
    if request.method == "POST":
        try:
            payload = json.loads(request.body.decode())
        except Exception:
            return JsonResponse({"error": "invalid JSON"}, status=400)

        action_type = payload.get("action_type")

        # ---------------------------
        # Approve / Reject / Pend user
        # ---------------------------
        if action_type == "update_user_status":
            if request.user.role not in ("admin", "superadmin"):
                return JsonResponse({"error": "forbidden"}, status=403)
            user_id = payload.get("user_id")
            new_status = payload.get("status")  # "approve"/"reject"/"pending"
            target_user = get_object_or_404(User, id=user_id)
            old_approved = target_user.approved
            if new_status == "approve":
                target_user.approved = True
            elif new_status == "reject":
                # we don't delete here automatically; mark approved False and optionally flag
                target_user.approved = False
            elif new_status == "pending":
                target_user.approved = False
            else:
                return JsonResponse({"error": "invalid status"}, status=400)
            target_user.save(update_fields=["approved"])
            # Log action
            ActionLog.objects.create(
                user=request.user,
                action_type=f"Updated user status -> {new_status}",
                model_name="User",
                object_id=str(target_user.id),
                details={"old_approved": old_approved, "new_approved": target_user.approved},
            )
            return JsonResponse({"ok": True, "message": f"{target_user.username} set to {new_status}"})

        # ---------------------------
        # Broadcast (admin can to teachers/students; teacher can to students only)
        # ---------------------------
        if action_type == "broadcast":
            role = payload.get("role")  # 'teacher' or 'student'
            message = payload.get("message", "").strip()
            if not role or not message:
                return JsonResponse({"error": "role & message required"}, status=400)
            # permission check: teacher can only send to 'student'
            if request.user.role == "teacher" and role != "student":
                return JsonResponse({"error": "forbidden"}, status=403)
            # admins can send to both teachers/students
            recipients = User.objects.filter(role=role, approved=True)
            created = 0
            for r in recipients:
                Notification.objects.create(sender=request.user, recipient=r, message=message, role=role, is_broadcast=True)
                created += 1
                ActionLog.objects.create(
                user=request.user,
                action_type="Broadcast",
                model_name="Notification",
                object_id="bulk",
                details={"role": role, "count": created, "message": message[:200]},
            )
            return JsonResponse({"ok": True, "message": f"Broadcast sent to {created} {role}(s).", "count": created})

    # -----------------------
    # GET: return dashboard data JSON
    # Supports pagination parameters: logs_page, quizzes_page, quizzes_page_size, logs_page_size
    # -----------------------
    # Stats
    stats = {
        "total_users": User.objects.filter(role__in=["admin", "teacher", "student"], approved=True).count(),
        "admins": User.objects.filter(role="admin", approved=True).count(),
        "teachers": User.objects.filter(role="teacher", approved=True).count(),
        "students": User.objects.filter(role="student", approved=True).count(),
        "pending_users": User.objects.filter(approved=False).count(),
        "classes": Class.objects.count(),
        "subjects": Subject.objects.count(),
        "quizzes": Quiz.objects.count(),
    }
    
    # Pending users (basic)
    pending_list_qs = User.objects.filter(approved=False).order_by("-date_joined")
    pending_list = list(pending_list_qs.values("id", "username", "email", "role", "date_joined")[:50])

    # Action logs paginated
    logs_page = int(request.GET.get("logs_page", 1))
    logs_page_size = int(request.GET.get("logs_page_size", 5))
    logs_qs = ActionLog.objects.order_by("-timestamp")
    paginator_logs = Paginator(logs_qs, logs_page_size)
    page_logs = paginator_logs.get_page(logs_page)
    logs = [
        {"action_type": l.action_type, "user": (l.user.username if l.user else "system"), "timestamp": l.timestamp.isoformat(), "details": l.details}
        for l in page_logs
    ]

        
    # Best student per class (leaderboard showing top student for each class)
    per_student_class_qs = (
        StudentQuizAttempt.objects.filter(is_submitted=True)
        .values(
            "student__id",
            "student__username",
            "student__first_name",
            "student__last_name",
            "student__student_class__id",
            "student__student_class__name",
        )
        .annotate(avg_score=Avg("score"))
        .order_by("student__student_class__id", "-score", "submitted_at")
    )

    # pick the top student per class (first row per class after ordering)
    best_in_class_map = {}
    for row in per_student_class_qs:
        cls_id = row.get("student__student_class__id") or 0
        if cls_id not in best_in_class_map:
            best_in_class_map[cls_id] = {
                "class_id": cls_id if cls_id != 0 else None,
                "class_name": row.get("student__student_class__name") or "Unknown",
                "student_id": row.get("student__id"),
                "username": row.get("student__username"),
                "first_name": row.get("student__first_name"),
                "last_name": row.get("student__last_name"),
                "avg_score": float(row.get("avg_score") or 0.0),
            }

   
    # final leaderboard: best student for each class
    leaderboard = list(best_in_class_map.values())

    # Class performance: average score per class (avg of attempts' scores grouped by student's class)
    class_perf_qs = (
        StudentQuizAttempt.objects.filter(is_submitted=True)
        .values("student__student_class__id", "student__student_class__name")
        .annotate(avg_score=Avg("score"))
        .order_by("student__student_class__name")
    )

    class_performance = [
        {
            "class_id": row.get("student__student_class__id") or None,
            "class_name": row.get("student__student_class__name") or "Unknown",
            "avg_score": float(row.get("avg_score") or 0.0),
        }
        for row in class_perf_qs
    
    ]

      # Show pending attempts initially (attempts that have at least one pending subjective answer)
    setter = User.objects.get(id=request.user.id, role__in=["admin", "superadmin"])
    pending_attempts = StudentQuizAttempt.objects.filter(quiz__created_by=setter, is_submitted=True).select_related('quiz', 'student').distinct().order_by('-submitted_at')
    grade_page_num = request.GET.get("grade_page", 1)

    grade_page = Paginator(pending_attempts, 10).get_page(grade_page_num)
   
    data = { "grading": [
    {
        "id": a.id,
        "student": a.student.username,
        "full_name": a.student.get_full_name(),
        "quiz": a.quiz.title,
        "quiz_id": a.quiz.id,
        "score": float(a.score or 0.0),
        "graded": bool(a.graded),
        "submitted_at": datetime.date(a.submitted_at) if getattr(a, "submitted_at", None) else None,
    }
    for a in grade_page
    ]
    }

    # Available quizzes (paginated) - show basic metadata
    quizzes_page = int(request.GET.get("quizzes_page", 1))
    quizzes_page_size = int(request.GET.get("quizzes_page_size", 10))
    quizzes_qs = Quiz.objects.select_related("subject", "created_by").order_by("-created_at")
    paginator_quiz = Paginator(quizzes_qs, quizzes_page_size)
    page_quiz = paginator_quiz.get_page(quizzes_page)
    quizzes = [
        {
            "id": q.id,
            "title": q.title,
            "subject": q.subject.name,
            "class_name": q.subject.school_class.name,
            "created_by": getattr(q.created_by, "username", str(q.created_by)),
            "start_time": q.start_time.isoformat(),
            "end_time": q.end_time.isoformat(),
            "is_published": q.is_published,
            "allow_retake": getattr(q, "allow_retake", False),
        }
        for q in page_quiz
    ]


    # Notifications for this admin (last 10)
    notifications_qs = Notification.objects.filter(recipient=request.user, is_read=False).order_by("-created_at")[:5]
    notifications = [{"id": n.id, "message": n.message, "sender": getattr(n.sender, "username", None), "created_at": n.created_at.isoformat()} for n in notifications_qs]
   

    return JsonResponse({
        "stats": stats,
        "pending_list": pending_list,
        "logs": logs,
        "logs_total_pages": paginator_logs.num_pages,
        "leaderboard": leaderboard,
        "class_performance": class_performance,
        "notifications": notifications,
        "quizzes": quizzes,
        "data": data,
        "quizzes_total_pages": paginator_quiz.num_pages,
        "grade_page_number": grade_page.number,
        "grade_num_pages": grade_page.paginator.num_pages,
        "paginator": grade_page.paginator.num_pages,
    })



from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone

@login_required
@user_passes_test(is_admin)
def api_admin_grading_page(request):
    """
    Returns JSON for admin grading table pagination.
    Query params:
      - page: page number (default 1)
      - page_size: rows per page (default 8)
      - q: optional search string to filter student or quiz
    """
    page = int(request.GET.get("page", 1))
    page_size = int(request.GET.get("page_size", 8))
    q = request.GET.get("q", "").strip()

    qs = StudentQuizAttempt.objects.select_related("student", "quiz").order_by("-submitted_at")

    # filter only submitted attempts (or change as required)
    qs = qs.filter(is_submitted=True)

    # simple search (student name, username or quiz title)
    if q:
        qs = qs.filter(
            Q(student__username__icontains=q) |
            Q(student__first_name__icontains=q) |
            Q(student__last_name__icontains=q) |
            Q(quiz__title__icontains=q)
        )

    paginator = Paginator(qs, page_size)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    # serialize attempts for JSON
    items = []
    for at in page_obj:
        student_name = at.student.get_full_name() if hasattr(at.student, "get_full_name") else getattr(at.student, "username", str(at.student))
        items.append({
            "id": at.id,
            "student_name": student_name,
            "quiz_title": at.quiz.title if at.quiz else "",
            "score": float(at.score) if hasattr(at, "score") and at.score is not None else (float(at.score) if getattr(at, "score", None) is not None else None),
            "graded": bool(getattr(at, "graded", False)),
            "submitted_at": at.submitted_at.isoformat() if at.submitted_at else None,
        })

    data = {
        "ok": True,
        "grading": items,
        "pagination": {
            "page": page_obj.number,
            "page_size": page_size,
            "num_pages": paginator.num_pages,
            "total": paginator.count,
            "has_next": page_obj.has_next(),
            "has_previous": page_obj.has_previous(),
            "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
            "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
        }
    }
    return JsonResponse(data)




@login_required
@user_passes_test(is_teacher_or_admin)
def admin_report_generator(request):
    """
    Display the report generator page where admin/teacher can choose class and date range.
    """
    classes = Class.objects.all() #prefetch_related(Prefetch('users', queryset=User.objects.filter(role='student')))
    return render(request, "exams/admin_report_generator.html", {"classes": classes})


import datetime
from django.http import FileResponse


### Maybe i can implement this later ------------------------------#
"""Generate PDF report for a single student over a date range."""

def admin_report_generator_for_single_student(request, student_id):
    student = get_object_or_404(User, pk=student_id)

    # 🕒 Date filtering - timezone aware
    start_date = timezone.make_aware(datetime.datetime(2025, 11, 1, 0, 0, 0))
    end_date = timezone.make_aware(datetime.datetime(2025, 11, 3, 23, 59, 59))

    attempts = (
        StudentQuizAttempt.objects.filter(student=student, is_submitted=True, submitted_at__range=(start_date, end_date))
        .select_related("quiz__subject")
        .order_by("-submitted_at")
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # 🎓 Header Section (Logo, School Name, Student Photo)
   
    school_logo = os.path.join(settings.BASE_DIR, 'static', 'images', 'school_logo2.png')

    if hasattr(student, 'profile_picture') and student.profile_picture:
        try:
            student_photo = Image(student.profile_picture.path, width=1.2*inch, height=1.2*inch)
        except Exception:
            student_photo = Paragraph("", styles['Normal'])
        else:
            student_photo = Paragraph("", styles['Normal'])

    table_data = []
    row = []
    if os.path.exists(school_logo):
        row.append(Image(school_logo, width=70, height=70))
    else:
        row.append("")

    row.append(Paragraph("<b><font size=16>Springfield High School</font></b>", styles["Title"]))

    if student_photo and os.path.exists(student_photo):
        row.append(Image(student_photo, width=70, height=70))
    else:
        row.append("")

    table_data.append(row)
    table = Table(table_data, colWidths=[80, 350, 80])
    table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    # 🧾 Student Info
    student_info = f"""
    <b>Name:</b> {student.first_name} {student.last_name}<br/>
    <b>Username:</b> {student.username}<br/>
    <b>Class:</b> {getattr(student.student_class, "name", "N/A")}<br/>
    <b>Total Attempts:</b> {attempts.count()}
    """
    elements.append(Paragraph(student_info, styles["Normal"]))
    elements.append(Spacer(1, 12))

    # 📊 Exam Table
    if not attempts.exists():
        elements.append(Paragraph("No quiz attempts found for this date range.", styles["Italic"]))
    else:
        data = [["Quiz Title", "Subject", "Score", "Date Submitted"]]
        for a in attempts:
            data.append([
                a.quiz.title,
                getattr(a.quiz.subject, "name", "N/A"),
                f"{a.score}%",
                timezone.localtime(a.submitted_at).strftime("%Y-%m-%d %H:%M"),
            ])

        t = Table(data, colWidths=[150, 100, 80, 120])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (2, 1), (2, -1), "CENTER"),
        ]))
        elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f"{student.username}_report.pdf")


def parse_date_safe(date_str):
    """
    Accepts either '2025-11-03' (from HTML input[type=date])
    or '03-11-2025' (manual input).
    """
    for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None 


@login_required
def admin_classes_report_pdf(request, class_id):
    from .models import Class
    ClassModel = Class
    class_obj = get_object_or_404(ClassModel, id=class_id)
    students = User.objects.filter(role='student', student_class=class_obj)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{class_obj.name}_report.pdf"'
    buffer = BytesIO()

    # 🧾 Minimal page margins
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20
    )
    styles = getSampleStyleSheet()
    elements = []

    # 🏫 Header info
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'school_logo2.png')
    school_name = getattr(settings, 'SCHOOL_NAME', "My School")
    school_address = getattr(settings, 'SCHOOL_ADDRESS', "123 School Street, City, Country")

    total_students = students.count()

    for student in students:
        # --- Header with logo, name, and student photo ---
        try:
            logo = Image(logo_path, width=1.0*inch, height=1.0*inch)
        except Exception:
            logo = Paragraph("", styles['Normal'])

        if getattr(student, 'profile_picture', None) and student.profile_picture:
            try:
                student_photo = Image(student.profile_picture.path, width=1.0*inch, height=1.0*inch)
            except Exception:
                student_photo = Paragraph("", styles['Normal'])
        else:
            student_photo = Paragraph("", styles['Normal'])

        header_data = [[
            logo,
            Paragraph(
                f"<b>{school_name}</b><br/><font size='9'>{school_address}</font>",
                ParagraphStyle('centered', fontSize=14, alignment=1)
            ),
            student_photo
        ]]

        header_table = Table(header_data, colWidths=[1.2*inch, 3.8*inch, 1.2*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 6))

        # --- Student info ---
        info_text = f"""
        <b>Student Name:</b> {student.get_full_name() or student.username}<br/>
        <b>Class:</b> {class_obj.name}<br/>
        <b>Total Students:</b> {total_students}
        """
        elements.append(Paragraph(info_text, ParagraphStyle('info', fontSize=10, leading=12)))
        elements.append(Spacer(1, 6))

        # --- Quiz Attempts ---
        attempts = StudentQuizAttempt.objects.filter(student=student, is_submitted=True).select_related('quiz__subject')

        if attempts.exists():
            subjects = [a.quiz.subject.name for a in attempts if a.quiz and a.quiz.subject]
            scores = [a.score for a in attempts if a.quiz and a.quiz.subject]

            # 🎯 Chart (score scale fixed 0–100, step 20)
            if subjects and scores:
                plt.figure(figsize=(5.3, 1.8))
                plt.bar(subjects, scores, color="#004c99")
                plt.title("Performance by Subject", fontsize=9, pad=5)
                plt.ylabel("Score (%)", fontsize=8)
                plt.ylim(0, 100)
                plt.yticks(range(0, 101, 20))
                plt.xticks(rotation=35, ha="right", fontsize=8)
                plt.tight_layout()
                chart_buf = BytesIO()
                plt.savefig(chart_buf, format='png', transparent=True)
                plt.close()
                chart_buf.seek(0)
                chart_img = Image(chart_buf, width=5.3*inch, height=1.8*inch)
                elements.append(chart_img)
                elements.append(Spacer(1, 6))

            # 📊 Table
            data = [["Exam Title", "Subject", "Score", "Date Submitted"]]
            for a in attempts:
                data.append([
                    a.quiz.title if a.quiz else "—",
                    a.quiz.subject.name if a.quiz and a.quiz.subject else "—",
                    f"{a.score:.1f}",
                    a.submitted_at.strftime("%d-%m-%Y %H:%M") if a.submitted_at else "—",
                ])

            table = Table(data, colWidths=[160, 120, 60, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#004c99")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("<i>No exam attempts found.</i>", styles['Normal']))

        # --- Footer Signature ---
        elements.append(Spacer(1, 10))
        footer = Paragraph(
            "__________________________<br/><b>Class Teacher / Admin Signature</b><br/>"
            f"<font size='8'>Generated on: {timezone.now().strftime('%d-%m-%Y %H:%M')}</font>",
            ParagraphStyle('footer', fontSize=8, alignment=1)
        )
        elements.append(footer)

        # Page break after each student
        elements.append(PageBreak())

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response



def parse_date_safe(date_str):
    """Accepts either 'YYYY-MM-DD' (HTML) or 'DD-MM-YYYY' (manual)."""
    if not date_str:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None



@login_required
@user_passes_test(lambda u: u.is_staff or u.role in ['admin', 'teacher'])
def admin_class_report_pdf(request):
    """
    Generate a compact one-page PDF report per student (with logo, address, photo, chart, and results table).
    """
    class_id = request.GET.get("class_id")
    start_date = parse_date_safe(request.GET.get("start_date"))
    end_date = parse_date_safe(request.GET.get("end_date"))

    # Validate class/date inputs
    if not class_id:
        return HttpResponse("Class not selected.", status=400)
    if not start_date or not end_date:
        return HttpResponse("Invalid or missing date format.", status=400)

    # Time zone aware and include the full last day
    start_date = timezone.make_aware(start_date)
    end_date = timezone.make_aware(end_date + timedelta(days=1))

    class_obj = get_object_or_404(Class, id=class_id)
    students = User.objects.filter(role="student", student_class=class_obj)

    # PDF setup
    response = HttpResponse(content_type='application/pdf')
    filename = f"{class_obj.name}_report_{timezone.now().strftime('%d-%m-%Y')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=30,
        rightMargin=30,
        topMargin=30,
        bottomMargin=25,
    )
    styles = getSampleStyleSheet()
    elements = []

    # Header info
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'school_logo2.png')
    school_name = getattr(settings, 'SCHOOL_NAME', "My School")
    school_address = getattr(settings, 'SCHOOL_ADDRESS', "123 School Street, City, Country")

    # Generate each student page
    for student in students:
        # School header
        try:
            logo = Image(logo_path, width=1.1 * inch, height=1.1 * inch)
        except Exception:
            logo = Paragraph("", styles['Normal'])

        # Student photo
        if getattr(student, 'profile_picture', None) and student.profile_picture:
            try:
                student_photo = Image(student.profile_picture.path, width=1.1 * inch, height=1.1 * inch)
            except Exception:
                student_photo = Paragraph("", styles['Normal'])
        else:
            student_photo = Paragraph("", styles['Normal'])

        header_data = [
            [
                logo,
                Paragraph(
                    f"<b>{school_name}</b><br/><font size='9'>{school_address}</font>",
                    ParagraphStyle('centered', fontSize=15, alignment=1),
                ),
                student_photo,
            ]
        ]
        header_table = Table(header_data, colWidths=[1.4 * inch, 3.6 * inch, 1.4 * inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 8))

        # Student info
        info = f"""
        <b>Student Name:</b> {student.get_full_name() or student.username}<br/>
        <b>Class:</b> {class_obj.name}<br/>
        <b>Total Students:</b> {students.count()}<br/>
        <b>Report Period:</b> {start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')}
        """
        elements.append(Paragraph(info, styles['Normal']))
        elements.append(Spacer(1, 8))

        # Quiz attempts in date range
        attempts = StudentQuizAttempt.objects.filter(
            student=student,
            is_submitted=True,
            submitted_at__range=(start_date, end_date)
        ).select_related('quiz', 'quiz__subject')

        if attempts.exists():
            subjects, scores = [], []

            for attempt in attempts:
                if attempt.quiz and attempt.quiz.subject:
                    subjects.append(attempt.quiz.subject.name)
                    scores.append(attempt.score)

            # Bar chart (score scale 0–100, step 20)
            if subjects and scores:
                plt.figure(figsize=(5.3, 2))
                plt.bar(subjects, scores, color="#004c99")
                plt.ylim(0, 100)
                plt.yticks(range(0, 101, 20))
                plt.ylabel("Score (%)", fontsize=9)
                plt.title("Performance by Subject", fontsize=10, pad=8)
                plt.xticks(rotation=40, ha="right", fontsize=8)
                plt.tight_layout()
                chart_buf = BytesIO()
                plt.savefig(chart_buf, format='png', transparent=True)
                plt.close()
                chart_buf.seek(0)
                chart_img = Image(chart_buf, width=5.3 * inch, height=2 * inch)
                elements.append(chart_img)
                elements.append(Spacer(1, 8))

            # Table of quiz results
            data = [["Exam Title", "Subject", "Score", "Date Submitted"]]
            for attempt in attempts:
                data.append([
                    attempt.quiz.title if attempt.quiz else "—",
                    attempt.quiz.subject.name if attempt.quiz and attempt.quiz.subject else "—",
                    f"{attempt.score:.1f}",
                    attempt.submitted_at.strftime("%d-%m-%Y %H:%M") if attempt.submitted_at else "—",
                ])

            table = Table(data, colWidths=[160, 120, 80, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#004c99")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 4),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("<i>No exam attempts found for this period.</i>", styles['Normal']))

        elements.append(Spacer(1, 10))

        # Signature area
        elements.append(Paragraph(
            "______________________________<br/><b>Class Teacher / Admin Signature</b><br/>"
            f"<font size='8'>Generated on: {timezone.now().strftime('%d-%m-%Y %H:%M')}</font>",
            ParagraphStyle('signature', fontSize=9, alignment=1)
        ))

        elements.append(PageBreak())

    # Build and return PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response



# ----------------- TEACHER DASHBOARD -------------------



def is_teacher(user):
    return user.is_authenticated and user.role == "teacher"


# ============================
# Teacher Dashboard
# ============================

@login_required
@user_passes_test(is_teacher_or_admin)

def teacher_dashboard(request):
    """Render teacher dashboard page"""
    teacher = request.user

    # students that have at least one submitted attempt in this teacher's class
    students_with_attempts_qs = User.objects.filter(
        studentquizattempt__is_submitted=True,
        studentquizattempt__student__student_class=teacher.student_class,
        role="student"
    ).distinct().order_by('last_name', 'first_name', 'username')

    # paginate students list for download/report links on dashboard
    students_page_num = int(request.GET.get("students_with_att_page", 1) or 1)
    students_per_page = 7  
    students_with_att_page = Paginator(students_with_attempts_qs, students_per_page).get_page(students_page_num)

    # pending subjective attempts for grading (existing behavior)
    pending_attempts = (
        StudentQuizAttempt.objects
        .filter(
            quiz__created_by=teacher,
            is_submitted=True,
            answers__is_pending=True,
            answers__question__question_type='subjective'
        )
        .select_related('quiz', 'student')
        .distinct()
        .order_by('-submitted_at')
    )

    return render(request, "exams/teacher_dashboard.html", {
        "attempts": pending_attempts,
        # keep both full queryset and paginated page (template can use page object)
        "students_with_attempts": students_with_attempts_qs,
        "students_with_att_page": students_with_att_page,
        "students_with_att_num": students_page_num,
        "students_with_att_num_pages": students_with_att_page.paginator.num_pages,
    })


@login_required
@user_passes_test(is_teacher_or_admin)
def teacher_dashboard_data(request):
    teacher = request.user
    student_class = getattr(teacher, "student_class", None)

    # Pagination
    quiz_page_num = request.GET.get("page", 1)
    notif_page_num = request.GET.get("notif_page", 1)
    grade_page_num = request.GET.get("grade_page", 1)
    broadcast_page_num = request.GET.get("broadcast_page", 1)

    # Attempts (all submitted attempts created by this teacher)
    attempts_qs = StudentQuizAttempt.objects.filter(
        quiz__created_by=teacher, is_submitted=True
    ).select_related('quiz', 'student').order_by('-submitted_at')
    
    # Quizzes
    quizzes = Quiz.objects.filter(created_by=teacher).annotate(
        attempt_count=Count("studentquizattempt")
    ).order_by("-created_at")

    quiz_page = Paginator(quizzes, 7).get_page(quiz_page_num)

    # Broadcasts
    broadcasts = Notification.objects.filter(sender=teacher).order_by("-created_at")
    broadcast_page = Paginator(broadcasts, 5).get_page(broadcast_page_num)

    # Notifications
    notifications = Notification.objects.filter(recipient=teacher).order_by("-created_at")
    notif_page = Paginator(notifications, 5).get_page(notif_page_num)

    # Grading: only submitted but NOT yet graded attempts (pending grading)
    teacher = request.user

    # Show pending attempts initially (attempts that have at least one pending subjective answer)
    pending_attempts = StudentQuizAttempt.objects.filter(quiz__created_by=teacher,is_submitted=True).select_related('quiz', 'student').distinct().order_by('-submitted_at')
    
    grade_page = Paginator(pending_attempts, 5).get_page(grade_page_num)
    

    # Performance
    performance = (
        StudentQuizAttempt.objects.filter(student__student_class=student_class)
        .values("student__username")
        .annotate(avg_score=Avg("score"))
    )

    data = {
        "summary": {
            "teacher": f" {teacher.username}, {teacher.first_name}",
            # return a simple count, not the QuerySet
            "attempts": attempts_qs.count(),
            "student_class": str(student_class),
            "total_quizzes": quizzes.count(),
            "my_quizzes": Quiz.objects.filter(created_by=teacher).count(),
            "my_class_quizzes": Quiz.objects.filter(created_by=teacher, school_class=student_class).count(),
            "objectives": Question.objects.filter(quiz__school_class=student_class, question_type="objective").count(),
            "pended":StudentQuizAttempt.objects.filter(quiz__created_by=teacher, is_submitted=True, answers__is_pending=True,answers__question__question_type='subjective').select_related('quiz', 'student').distinct().order_by('-submitted_at').count(),
            "graded": Answer.objects.filter(question__quiz__created_by=teacher, is_pending=False).count(),
            "attempts_total": StudentQuizAttempt.objects.filter(student__student_class=student_class).count(),
            "total_students": User.objects.filter(student_class=student_class, role="student").count(),
        },
        "quizzes": [
            {
                "id": q.id,
                "title": q.title,
                "subject": q.subject.name,
                "class_name": q.subject.school_class.name,
                "created_at": datetime.date(q.created_at),
                "end_time": datetime.date(q.end_time),
                "attempts": q.attempt_count,
                "allow_retake": getattr(q, "allow_retake", False),
            }
            for q in quiz_page
        ],
        "notifications": [
            {
                "id": n.id,
                "message": n.message,
                "is_read": n.is_read,
                "created_at": datetime.date(n.created_at),
            }
            for n in notif_page
        ],
        "broadcasts": [
            {
                "message": b.message,
                "get_target_display": b.get_target_display(),
                "created_at": datetime.date(b.created_at),
            }
            for b in broadcast_page
        ],
        "performance": list(performance),
       
        "grading": [
            {
                "id": a.id,
                "student": a.student.username,
                "full_name": a.student.get_full_name(),
                "quiz": a.quiz.title,
                "quiz_id": a.quiz.id,
                "score": float(a.score or 0.0),
                "graded": bool(a.graded),
                "submitted_at": datetime.date(a.submitted_at) if getattr(a, "submitted_at", None) else None,
            }
            for a in grade_page
        ],

        "pagination": {
            "quiz_page_number": quiz_page.number,
            "quiz_num_pages": quiz_page.paginator.num_pages,
            "notif_page_number": notif_page.number,
            "notif_num_pages": notif_page.paginator.num_pages,
            "broadcast_page_number": broadcast_page.number,
            "broadcast_num_pages": broadcast_page.paginator.num_pages,
            "grade_page_number": grade_page.number,
            "grade_num_pages": grade_page.paginator.num_pages,
        },
    }
   
    return JsonResponse(data)


# ============================
# Teacher Broadcast
# ============================
@login_required
@user_passes_test(is_teacher)
def teacher_broadcast(request):
    """Teacher sends a broadcast"""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=400)

    message = request.POST.get("message")
    audience = request.POST.get("audience")  # "students" | "admin"

    if not message or not audience:
        return JsonResponse({"error": "Missing fields"}, status=400)

    if audience == "students":
        recipients = User.objects.filter(role="student")
    elif audience == "admin":
        recipients = User.objects.filter(role="admin")
    else:
        return JsonResponse({"error": "Invalid audience"}, status=400)

    for r in recipients:
        Notification.objects.create(sender=request.user, recipient=r, message=message)

    ActionLog.objects.create(
        user=request.user,
        action_type="Teacher Broadcast",
        model_name="Notification",
        object_id=str(request.user.id),
        details={"message": message, "audience": audience},
    )

    return JsonResponse({"success": True, "message": f"Broadcast sent to {audience}."})


# ============================
# Grading Endpoint
# ============================

@login_required
@user_passes_test(is_teacher)
def grading_list(request):
    ungraded_answers = Answer.objects.filter(is_pending=True).select_related("attempt", "question__quiz")
    attempt = {}
    for a in ungraded_answers:
        attempt.update({'id': a.attempt.id, 'quiz': a.attempt.quiz.id})
   
    return render(request, "exams/grading_list.html", {"answers": ungraded_answers, 'attempt': attempt})


@login_required
@user_passes_test(is_teacher_or_admin)
def grade_attempt(request, attempt_id):
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id)
    answers = Answer.objects.filter(attempt=attempt, question__question_type='subjective')

    # helper to robustly detect AJAX
    def _is_ajax(req):
        return req.headers.get('X-Requested-With') == 'XMLHttpRequest' or req.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

    if request.method == "POST":
        total_subjective = 0
        for ans in answers:
            score = request.POST.get(f'score_{ans.id}', '0')
            try:
                ans.score = float(score)
                ans.is_pending = False
                ans.graded_by = request.user
                ans.graded_at = timezone.now()
                ans.save(update_fields=["score", "is_pending", "graded_by", "graded_at"])
                total_subjective += ans.score
            except ValueError:
                continue

        total_objective = Answer.objects.filter(
            attempt=attempt, question__question_type='objective'
        ).aggregate(Sum('score'))['score__sum'] or 0.0


        total_score = total_objective + total_subjective
        attempt.score = total_score
        attempt.graded = True
        attempt.save(update_fields=["score", "graded"])

        if _is_ajax(request):
            return JsonResponse({
                "success": True,
                "message": f"Grading completed. Total score = {total_score}",
                "score": total_score
            })

        messages.success(request, "✅ Grading completed successfully!")
        return redirect("users:teacher_dashboard")

    # If modal load (AJAX)
    if _is_ajax(request):
        return render(request, "exams/partials/grade_form.html", {
            "attempt": attempt,
            "answers": answers,
        })
    
    ActionLog(
        user = request.user,
        action_type = 'Grade',
        description = 'grading exam',
        created_at = timezone.now(),
        model_name = Answer,
        object_id = str(ans.quiz.id),
        timestamp = timezone.now(),
        details = "Your exam has been graded"
    )

    Notification(
        sender = request.user,
        recipient = get_object_or_404(User, id=attempt_id, role="student"),
        message=f"Your exam has been graded: {ans.quiz.title}",
        created_at = timezone.now()
    )

    return render(request, "exams/grading_attempt.html", {
        "attempt": attempt,
        "answers": answers,
    })


# ============================
# Student Report PDF Generation


@login_required
def student_report_pdf(request, student_id):
    student = get_object_or_404(User, id=student_id, role='student')
    attempts = StudentQuizAttempt.objects.filter(student=student).select_related('quiz__subject')

    # Prepare PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{student.username}_report.pdf"'
    buffer = BytesIO()

    # PDF setup
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    # =============================
    # HEADER SECTION (LOGO + TITLE + PHOTO)
    # =============================
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'school_logo2.png') 
    school_name = settings.SCHOOL_NAME if hasattr(settings, 'SCHOOL_NAME') else "My School"

    # Try loading logo
    try:
        logo = Image(logo_path, width=1.2*inch, height=1.2*inch)
    except Exception:
        logo = Paragraph("", styles['Normal'])

    # Try loading student photo (if exists)
    if hasattr(student, 'profile_picture') and student.profile_picture:
        photo_path = student.profile_picture.path
        try:
            student_photo = Image(photo_path, width=1.2*inch, height=1.2*inch)
        except Exception:
            student_photo = Paragraph("", styles['Normal'])
    else:
        student_photo = Paragraph("", styles['Normal'])

    # Header layout (logo, school name, student photo)
    header_data = [
        [logo, Paragraph(f"<b>{school_name}</b>", ParagraphStyle('centered', fontSize=16, alignment=1)), student_photo]
    ]
    header_table = Table(header_data, colWidths=[1.5*inch, 3.5*inch, 1.5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # =============================
    # STUDENT INFO SECTION
    # =============================
    total_students = User.objects.filter(role='student', student_class=student.student_class).count()

    info_text = f"""
    <b>Student Name:</b> {student.get_full_name() or student.username}<br/>
    <b>Class:</b> {student.student_class.name if hasattr(student, 'student_class') else 'N/A'}<br/>
    <b>Total Students in Class:</b> {total_students}
    """
    elements.append(Paragraph(info_text, styles['Normal']))
    elements.append(Spacer(1, 12))

    # =============================
    # PERFORMANCE CHART (MATPLOTLIB)
    # =============================
    if attempts.exists():
        subjects = []
        scores = []

        for attempt in attempts:
            if attempt.quiz.subject:
                subjects.append(attempt.quiz.subject.name)
                scores.append(attempt.score or 0.0)

        if subjects:
            fig = Figure(figsize=(6, 2))
            ax = fig.add_subplot(111)
            ax.bar(subjects, scores, color='#2a9df4')
            ax.set_title("Student Performance by Subject")
            ax.set_xlabel("Subject")
            ax.set_ylabel("Score")
            ax.tick_params(axis='x', rotation=45)
            fig.tight_layout()

            canvas = FigureCanvas(fig)
            chart_buffer = BytesIO()
            canvas.print_png(chart_buffer)
            chart_buffer.seek(0)

            chart_image = Image(chart_buffer, width=5.5*inch, height=4.5*inch)
            elements.append(chart_image)
            elements.append(Spacer(1, 20))

    # =============================
    # EXAM ATTEMPT TABLE
    # =============================
    elements.append(Paragraph("<b>Exam Attempt Details</b>", styles['Heading3']))
    elements.append(Spacer(1, 8))

    if not attempts.exists():
        elements.append(Paragraph("No Exam attempts yet.", styles['Normal']))
    else:
        data = [["Exam Title", "Subject", "Score", "Date Taken"]]
        for attempt in attempts:
            data.append([
                attempt.quiz.title,
                attempt.quiz.subject.name if attempt.quiz.subject else "—",
                f"{attempt.score}",
                attempt.submitted_at.strftime("%d-%m-%Y %H:%M") if attempt.submitted_at else "—"
            ])

        table = Table(data, colWidths=[160, 120, 80, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#003366")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ]))
        elements.append(table)

    # =============================
    # BUILD DOCUMENT
    # =============================
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response



# ----------------- STUDENT REVIEW  on teacher dashboard-------------------
@login_required
@user_passes_test(is_teacher)
def student_review(request, student_id):
    student = get_object_or_404(User, id=student_id, role="student"),
    attempts = StudentQuizAttempt.objects.filter(student=student).select_related("quiz"),
    return render(request, "exam/quick_review.html", {"student": student, "attempts": attempts})



# ----------------- RETAKE REQUEST -------------------
@login_required
@user_passes_test(is_teacher)
def approve_retake(request, quiz_id, student_id):
    student = get_object_or_404(User, id=student_id, role="student")
    quiz = get_object_or_404(Quiz, id=quiz_id, created_by=request.user)

    attempt, _ = StudentQuizAttempt.objects.get_or_create(student=student, quiz=quiz)
    attempt.retake_allowed = True
    attempt.is_submitted = False
    attempt.end_time = None
    attempt.retake_count += 1
    attempt.save()
  
    Notification.objects.create(user=student, recipient=student, message=f"You can now retake Exam: {quiz.title}")
    ActionLog.objects.create(user=request.user, action_type="Approved Retake", model_name="Exam", object_id=str(quiz.id))

    return JsonResponse({"success": True})

# ----------------- BROADCAST MESSAGE -------------------
@login_required
@user_passes_test(is_teacher)
def broadcast_message(request):
    """
    Allows a teacher to send broadcast messages to students in their classes or admins.
    """
    if request.method == "POST":
        message = request.POST.get("message", "").strip()
        audience = request.POST.get("audience", "students")

        if not message:
            return JsonResponse({"error": "Message cannot be empty."}, status=400)

        if audience == "students":
            # Send to students in teacher's class
            students = User.objects.filter(
                role="student", school_class__in=request.user.classes.all()
            )
        elif audience == "admin":
            students = User.objects.filter(role="admin")
        else:
            return JsonResponse({"error": "Invalid audience."}, status=400)

        for student in students:
            Notification.objects.create(
                recipient=student,
                sender=request.user,
                message=message,
                is_broadcast=True,
                role="student" if audience == "students" else "admin",
                created_at=timezone.now()
            )

        ActionLog.objects.create(
            user=request.user,
            action="Broadcast Message",
            model_name="Notification",
            object_id=str(request.user.id)
        )

        return JsonResponse({"success": True, "message": "Broadcast sent successfully!"})
    return JsonResponse({"error": "Invalid request method."}, status=400)

# ----------------- MARK NOTIFICATION READ -------------------
@login_required
def mark_notification_read(request, notification_id):
    """
    Marks a specific notification as read.
    """
    notif = get_object_or_404(Notification, id=notification_id, recipient=request.user)
    notif.is_read = True
    notif.save()
    return JsonResponse({"success": True})

# ----------------- DOWNLOAD REPORTS -------------------
@login_required
@user_passes_test(is_teacher)
def download_student_report(request, student_id):
    student = get_object_or_404(User, id=student_id, role="student")
    attempts = StudentQuizAttempt.objects.filter(student=student)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)
    story = []
    styles = getSampleStyleSheet()

    story.append(Paragraph(f"<b>School Name</b>", styles["Title"]))
    story.append(Paragraph(f"Student: {student.get_full_name()}", styles["Heading2"]))
    story.append(Spacer(1, 12))

    for attempt in attempts:
        story.append(Paragraph(f"Exam: {attempt.quiz.title} | Date: {attempt.start_time}", styles["Normal"]))
        for ans in attempt.answer.all():
            story.append(Paragraph(f"- {ans.question.text}: {ans.score}", styles["Normal"]))

    doc.build(story)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{student.username}_report.pdf"'
    return response


@login_required
@user_passes_test(is_teacher)
def download_quiz_report(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, created_by=request.user)
    attempts = StudentQuizAttempt.objects.filter(quiz=quiz)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)
    story = []
    styles = getSampleStyleSheet()

    story.append(Paragraph(f"<b>School Name</b>", styles["Title"]))
    story.append(Paragraph(f"Exam Report: {quiz.title}", styles["Heading2"]))
    story.append(Spacer(1, 12))

    for attempt in attempts:
        student = attempt.student
        story.append(Paragraph(f"Student: {student.get_full_name()} | Date: {attempt.start_time}", styles["Normal"]))
        for ans in attempt.answer.all():
            story.append(Paragraph(f"- {ans.question.text}: {ans.score}", styles["Normal"]))
        story.append(Spacer(1, 6))

    doc.build(story)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="quiz_{quiz.id}_report.pdf"'
    return response

# ----------------------Teacher dashboard data endpoint (JSON) (for AJAX refresh)-------------------------#



###--------------------------------Student Dashboard -------------------------------###

def is_student(user):
    return user.is_authenticated and user.role == 'student' 


@login_required
@user_passes_test(is_student)
def student_dashboard(request):
    """Render the student dashboard shell. Data comes from student_dashboard_data (AJAX)."""
    return render(request, "exams/student_dashboard.html")


@login_required
@user_passes_test(is_student)
def student_dashboard_data(request):
    """
    Returns JSON containing:
      - notifications (paginated)
      - summary (counts)
      - available_quizzes (paginated)
      - past_attempts (paginated)
      - leaderboard (top 10)
      - performance_chart (subject -> %)
    Query params (optional):
      notif_page, notif_page_size, quizzes_page, quizzes_page_size, attempts_page, attempts_page_size
    """
    student = request.user
    student_class = getattr(student, "student_class", None)

    # pagination params
    notif_page = int(request.GET.get("notif_page", 1))
    notif_page_size = int(request.GET.get("notif_page_size", 5))
    quizzes_page = int(request.GET.get("quizzes_page", 1))
    quizzes_page_size = int(request.GET.get("quizzes_page_size", 5))
    attempts_page = int(request.GET.get("attempts_page", 1))
    attempts_page_size = int(request.GET.get("attempts_page_size", 5))

    # ---------------- Notifications (paginated) ----------------
    notif_qs = Notification.objects.filter(recipient=student).order_by("-created_at")
    notif_p = Paginator(notif_qs, notif_page_size)
    notif_page_obj = notif_p.get_page(notif_page)
    notifications = [
        {"id": n.id, "message": n.message, "is_read": n.is_read, "created_at": datetime.date(n.created_at)}
        for n in notif_page_obj
    ]
    notif_meta = {"page": notif_page_obj.number, "pages": notif_p.num_pages, "total": notif_p.count}

    
    # ---------------- Available quizzes (paginated) ----------------

    now = timezone.now()
    if student_class is None:
        available_qs = Quiz.objects.none()
    else:
        available_qs = Quiz.objects.filter(
        school_class__name=student_class,
        is_published=True, 
        end_time__gte=now 
        ).order_by("-created_at")

        # exclude quizzes already completed without retake allowed
        exclude_ids = StudentQuizAttempt.objects.filter(
            student=student, is_submitted=True, retake_allowed=False,
        ).values_list("quiz_id", flat=True)
        available_qs = available_qs.exclude(id__in=exclude_ids) 

    qp = Paginator(available_qs, quizzes_page_size)
    qp_obj = qp.get_page(quizzes_page)

    quizzes_data = []
    # build rich data (can't easily get allow_retake from values; use getattr)
    for q in qp_obj:
        # check last attempt status
        last_attempt = StudentQuizAttempt.objects.filter(student=student, quiz=q).order_by("-started_at").first()
        already_submitted = StudentQuizAttempt.objects.filter(student=student, quiz=q, is_submitted=True).exists()
        student_retake_override = bool(last_attempt and getattr(last_attempt, "retake_allowed", False))
        allow_retake_global = getattr(q, "allow_retake", False) if hasattr(q, "allow_retake") else False
        
        # Added lately for great UI/UX. 
        latest_request = RetakeRequest.objects.filter(student=request.user, quiz=q).last()
        retake_request_count = RetakeRequest.objects.filter(student=already_submitted, quiz=q).count()
        retake_status = latest_request.status if latest_request else None
        
        quizzes_data.append({
            "id": q.id,
            "title": q.title,
            "subject": q.subject.name if q.subject else "",
            "class_name": q.subject.school_class.name if (q.subject and q.subject.school_class) else "",
            "start_time": datetime.date(q.start_time) if q.start_time else None,
            "end_time": datetime.date(q.end_time) if q.end_time else None,
            "duration_minutes": getattr(q, "duration_minutes", None),
            "is_published": bool(q.is_published),
            "allow_retake": bool(allow_retake_global),
            "already_submitted": bool(already_submitted),
            "student_retake_override": student_retake_override,
            "retake_status": retake_status,  # 🔹 added
        })
            
    quizzes_meta = {"page": qp_obj.number, "pages": qp.num_pages, "total": qp.count}

 
      # ---------------- Summary ----------------

    total_attempts = StudentQuizAttempt.objects.filter(student=student).count()
    auto_graded_count = Answer.objects.filter(attempt__student=student, is_pending=False).count()
    pending_subjectives = Answer.objects.filter(attempt__student=student, is_pending=True).count()

    summary = {
        "total_attempts": total_attempts,
        "auto_graded_count": auto_graded_count,
        "pending_subjectives": pending_subjectives,
    } 


    # ---------------- Past attempts (paginated) ----------------
    attempts_qs = StudentQuizAttempt.objects.filter(student=student, is_submitted=True, retake_requested=False).select_related("quiz").order_by("-started_at")
    ap = Paginator(attempts_qs, attempts_page_size)
    ap_obj = ap.get_page(attempts_page)

    past_attempts = []
    for a in ap_obj:
        # compute totals for this attempt using Answer model
        answer = Answer.objects.filter(attempt=a).select_related("question", "selected_choice")
        total_score = Answer.objects.filter(attempt=a).aggregate(total=Sum('score'))['total'] or 0
        total_qmarks = 0.0
        obtained = 0.0
        pending_subjectives_count = 0
        wrong_review = []
        for ans in answer:
            q_marks = getattr(ans.question, "marks", 0) or 0
            total_qmarks += q_marks
            if ans.is_pending:
                pending_subjectives_count += 1
                # subjective might not have marks yet; don't include
            else:
                # for objective and graded subjective
                obtained += float(ans.score or 0.0)
            # for review: detect wrong objective answer
            # if question is objective and selected_choice exists but is incorrect
            if getattr(ans.question, "question_type", getattr(ans.question, "type", None)) in ("objective", "multiple_choice"):
                if not (ans.selected_choice and getattr(ans.selected_choice, "is_correct", False)):
                    # get correct answer text(s)
                    correct_choices = ans.question.choices.filter(is_correct=True).values_list("text", flat=True)
                    correct_text = ", ".join(correct_choices) if correct_choices else ""
                    user_ans = ans.selected_choice.text if ans.selected_choice else "-"
                    wrong_review.append({
                        "question": ans.question.text,
                        "your_answer": user_ans,
                        "correct_answer": correct_text
                    })

        past_attempts.append({
            "attempt_id": a.id,
            "quiz_id": a.quiz.id,
            "quiz": a.quiz.title if a.quiz else "", 
            "subject": a.quiz.subject.name,
            "total_score": total_score,
            "obtained": obtained,
            "total_qmarks": total_qmarks,
            "pending_subjectives": pending_subjectives_count,
            "is_submitted": bool(a.is_submitted),
            "started_at": datetime.date(a.started_at) if getattr(a, "started_at", None) else None,
            "submitted_at": datetime.date(getattr(a, "submitted_at", None)) if getattr(a, "submitted_at", None) else None,
            "wrong_answer": wrong_review,
            "retake_count": getattr(a, "retake_count", 0),
        })

    
    attempts_meta = {"page": ap_obj.number, "pages": ap.num_pages, "total": ap.count}


  
    # ---------------- Leaderboard (top 10 students) ----------------
    lb_qs = (
        StudentQuizAttempt.objects.filter(is_submitted=True, student__student_class=request.user.student_class)
        .values("student__id", "student__username", "student__first_name", "student__last_name")
        .annotate(avg_score=Avg("score"))
        .order_by("-avg_score")[:10]
    )
    leaderboard = [
        {
            "student_id": r["student__id"],
            "username": r.get("student__username"),
            "full_name": f"{r.get('student__first_name') or ''} {r.get('student__last_name') or ''}".strip(),
            "avg_score": float(r["avg_score"] or 0)
        }
        for r in lb_qs
    ]

    # ---------------- Performance Chart (per subject in student's class) ----------------
    

    attempts = StudentQuizAttempt.objects.filter(student=student, is_submitted=True)

    performance_chart = []
    for att in attempts:    
        perf_qs = Answer.objects.filter(
            attempt=att,
            attempt__quiz__school_class=student_class
        ).values("attempt__quiz__subject__name").annotate(
            obtained=Sum('score'),
            possible=Sum("question__marks")
        )
        for r in perf_qs:
            subj = r.get("attempt__quiz__subject__name") or "Unknown"
            obtained = float(r.get("obtained") or 0)
            possible = float(r.get("possible") or 0) or 0.0
            pct = round((obtained / possible) * 100, 2) if possible > 0 else 0.0
            performance_chart.append({"subject": subj, "obtained": obtained, "possible": possible, "percentage": pct})
   

    # ---------------- Return JSON ----------------

    return JsonResponse({
        "notifications": notifications,
        "notifications_meta": notif_meta,
        "summary": summary,
        "available_quizzes": quizzes_data,
        "available_quizzes_meta": quizzes_meta,
        "past_attempts": past_attempts,
        "past_attempts_meta": attempts_meta,
        "leaderboard": leaderboard,
        "performance_chart": performance_chart,
    })




#------------------------------------------------------------------------
@require_POST
@login_required
@user_passes_test(is_student)
def api_notifications_mark_read(request):
    """Mark notification read (AJAX POST: {id: notification_id})"""
    import json
    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "invalid json"}, status=400)

    nid = payload.get("id")
    if not nid:
        return JsonResponse({"ok": False, "error": "id required"}, status=400)

    notif = get_object_or_404(Notification, id=nid, recipient=request.user)
    notif.is_read = True
    notif.save(update_fields=["is_read"])
    return JsonResponse({"ok": True, "id": nid})


### ------------------------------ Student_dashboard Ended ----------------------------------------------------##

def _is_admin(user):
    return user.is_authenticated and user.role in ("admin", "superadmin")

def _is_teacher(user):
    return user.is_authenticated and user.role == "teacher"


# Broadcast endpoint
# -------------------
@login_required
@require_POST
def api_broadcast(request):
    """
    POST JSON { "role": "student"|"teacher", "message": "..." }
    - Admins (admin/superadmin) may broadcast to 'teacher' or 'student'
    - Teachers may broadcast only to 'student'
    Creates Notification objects for each recipient and logs ActionLog.
    """
    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    role = payload.get("role")
    message = (payload.get("message") or "").strip()
    if not role or not message:
        return JsonResponse({"ok": False, "error": "role and message required"}, status=400)

    # Permissions: teacher can only send to students
    if request.user.role == "teacher" and role != "student":
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    # Only allow admin/teacher to broadcast
    if request.user.role not in ("admin", "superadmin", "teacher"):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    recipients = User.objects.filter(role=role, approved=True)
    created = 0
    for r in recipients:
        # Notification.objects.create(sender=request.user, recipient=r, role=role, message=message, created_at=timezone.now())
        created += 1

    ActionLog.objects.create(
        user=request.user,
        action="Broadcast sent",
        model_name="Notification",
        object_id="bulk",
        details={"role": role, "count": created, "sample": message[:120]},
        timestamp=timezone.now()
    )

    return JsonResponse({"ok": True, "message": f"Broadcast sent to {created} {role}(s).", "count": created})


# -------------------
# Get unread notifications for current user
# -------------------
@login_required
def api_notifications_unread(request):
    """
    GET -> returns unread (is_read=False) notifications for request.user
    """
    qs = Notification.objects.filter(recipient=request.user, is_read=False).order_by("-created_at")
    data = [
        {
            "id": n.id,
            "message": n.message,
            "sender": n.sender.username if n.sender else None,
            "created_at": n.created_at.isoformat()
        }
        for n in qs
    ]
    return JsonResponse({"ok": True, "notifications": data})


# -------------------
# Mark notification as read
# -------------------

@login_required
@require_POST
def api_notifications_mark_read(request):
    """
    # POST JSON { "id": <notification_id> } -> set is_read = True for that notification if it belongs to current user
    """
    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    nid = payload.get("id")
    if not nid:
        return JsonResponse({"ok": False, "error": "id required"}, status=400)

    notif = get_object_or_404(Notification, id=nid, recipient=request.user)
    notif.is_read = True
    notif.save(update_fields=["is_read"])

    ActionLog.objects.create(
        user=request.user,
        action="Read notification",
        model_name="Notification",
        object_id=str(nid),
        # details={"message_sample": (notif.message[:80] if notif.message else "")},
        timestamp=timezone.now()
    )

    return JsonResponse({"ok": True, "message": "marked read", "id": nid})


####--------------------------No 1 Endpoint Ended----------------------------####


@require_http_methods(["GET"])
def api_notifications_unread(request):
    """Return unread notifications for current user."""
    if not request.user.is_authenticated:
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    from users.models import Notification  # import here to avoid circular issues
    qs = Notification.objects.filter(recipient=request.user, is_read=False).order_by("-created_at")
    notifs = [{
        "id": n.id,
        "message": n.message,
        "sender": n.sender.username if n.sender else None,
        "created_at": n.created_at.isoformat()
    } for n in qs]
    return JsonResponse({"ok": True, "notifications": notifs})


@require_http_methods(["POST"])
def api_notifications_mark_read(request):
    """Mark notification as read (student/teacher/admin)."""
    if not request.user.is_authenticated:
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "invalid json"}, status=400)

    nid = payload.get("id")
    if not nid:
        return JsonResponse({"ok": False, "error": "id required"}, status=400)

    from users.models import Notification
    notif = get_object_or_404(Notification, id=nid, recipient=request.user)
    notif.is_read = True
    notif.save(update_fields=["is_read"])
    ActionLog.objects.create(user=request.user, action_type="Read notification", model_name="Notification", object_id=str(nid))
    return JsonResponse({"ok": True, "id": nid})



# old API -------------------------------------------------#


# -------------------------- No 2 Endpoint Ended -------------------------- #



# ---- USER MANAGEMENT ----
@login_required
@user_passes_test(is_admin)
def manage_users(request):
    users = User.objects.exclude(role="superadmin").order_by("-date_joined")
    return render(request, "exams/manage_users.html", {"users": users})


@login_required
@user_passes_test(is_admin)
def create_user(request):
    if request.method == "POST":
        username = request.POST["username"]
        email = request.POST["email"]
        role = request.POST["role"]
        password = request.POST["password"]

        user = User.objects.create_user(username=username, email=email, password=password, role=role, approved=True)
        ActionLog.objects.create(user=request.user, action=f"Created user {user.username} ({role})")
        messages.success(request, f"User {username} created successfully.")
        return redirect("users:manage_users")

    return render(request, "exams/create_user.html")



# -------------------- CLASSES --------------------
def manage_classes_subjects(request):
    return render(request, 'exams/class_subject.html')

def serialize_all():
    """Return all classes and subjects in one dict"""
    classes = list(Class.objects.values("id", "name"))
    subjects = list(Subject.objects.values("id", "name", "school_class_id"))
    return {"classes": classes, "subjects": subjects}

@csrf_exempt
def class_subject_crud(request):
    if request.method == "GET":
        return JsonResponse(serialize_all(), safe=False)

    if request.method == "POST":
        data = json.loads(request.body)
        action = data.get("action")

        # ✅ CREATE CLASS
        if action == "create_class":
            Class.objects.create(name=data["name"])

        # ✅ UPDATE CLASS
        elif action == "update_class":
            try:
                cls = Class.objects.get(pk=data["id"])
                cls.name = data["name"]
                cls.save()
            except Class.DoesNotExist:
                return JsonResponse({"error": "Class not found"}, status=404)

        # ✅ DELETE CLASS
        elif action == "delete_class":
            Class.objects.filter(pk=data["id"]).delete()

        # ✅ CREATE SUBJECT
        elif action == "create_subject":
            Subject.objects.create(name=data["name"], school_class_id=data["class_id"])

        # ✅ UPDATE SUBJECT
        elif action == "update_subject":
            try:
                subj = Subject.objects.get(pk=data["id"])
                subj.name = data["name"]
                subj.school_class_id = data["class_id"]
                subj.save()
            except Subject.DoesNotExist:
                return JsonResponse({"error": "Subject not found"}, status=404)

        # ✅ DELETE SUBJECT
        elif action == "delete_subject":
            Subject.objects.filter(pk=data["id"]).delete()

        return JsonResponse(serialize_all(), safe=False)

    return JsonResponse({"error": "Invalid request"}, status=400)



# ---- QUIZ MANAGEMENT ----
@login_required
@user_passes_test(is_admin)
def manage_quizzes_admin(request):
    quizzes = Quiz.objects.all().order_by("-created_at")

    page = Paginator(quizzes, 7)
    page_number = request.GET.get("page")
    page_obj = page.get_page(page_number)

    return render(request, "exams/manage_quizzes.html", {"page_obj": page_obj})

# @login_required
# @user_passes_test(is_teacher_or_admin)
def manage_quizzes_teacher(request):
    quizzes = Quiz.objects.filter(created_by=request.user).order_by("-created_at")

    page = Paginator(quizzes, 7)
    page_number = request.GET.get("page")
    page_obj = page.get_page(page_number)

    return render(request, "exams/manage_quizzes_teacher.html", {"page_obj": page_obj})


# @login_required
# @user_passes_test(is_teacher_or_admin)
def search_quizzes(request):
    q = request.GET.get("q", "")
    page_number = request.GET.get("page", 1)

    quizzes = Quiz.objects.all()
    if q:
        quizzes = quizzes.filter(
            title__icontains=q
        ) | quizzes.filter(
            subject__name__icontains=q
        ) | quizzes.filter(
            subject__school_class__name__icontains=q
        )

    paginator = Paginator(quizzes.order_by("-start_time"), 5)  # 5 per page
    page_obj = paginator.get_page(page_number)

    data = {
        "results": [
            {
                "id": quiz.id,
                "title": quiz.title,
                "subject": quiz.subject.name,
                "class_name": quiz.subject.school_class.name,
                "start_time": datetime.date(quiz.start_time),
                "end_time": datetime.date(quiz.end_time),
                "is_published": quiz.is_published,
            }
            for quiz in page_obj
        ],
        "current_page": page_obj.number,
        "num_pages": paginator.num_pages,
        "has_previous": page_obj.has_previous(),
        "has_next": page_obj.has_next(),
        "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
        "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
    }
    return JsonResponse(data)



@login_required
@user_passes_test(is_teacher_or_admin)
def create_quiz(request):
    if request.method == "POST":
        title = request.POST["title"]
        subject_id = request.POST["subject"]
        duration = int(request.POST["duration"])

        subject = Subject.objects.get(id=subject_id)
        quiz = Quiz.objects.create(title=title, subject=subject, duration=duration, created_by=request.user)
        ActionLog.objects.create(user=request.user, action_type=f"Created quiz {quiz.title}")
        messages.success(request, f"Quiz {title} created successfully.")
        return redirect("exams:manage_quizzes")

    subjects = Subject.objects.all()
    return render(request, "exams/create_quiz.html", {"subjects": subjects})


@login_required
@user_passes_test(is_admin)
def upload_quiz_excel(request):
    if request.method == "POST" and request.FILES["file"]:
        file = request.FILES["file"]
        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        filepath = fs.path(filename)

        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header
            title, subject_name, duration = row[:3]
            subject, _ = Subject.objects.get_or_create(name=subject_name)
            Quiz.objects.create(title=title, subject=subject, duration=int(duration), created_by=request.user)

        ActionLog.objects.create(user=request.user, action_type=f"Uploaded Exams from Excel")
        messages.success(request, "Exams uploaded successfully.")
        return ("exams:manage_quizzes")

    return render(request, "exams/upload_quiz_excel.html")


# -----------------------------Student Quiz List with Status---------------------------------#  
from django.db.models import Q  

def get_quizzes_with_status(student):
    # Fetch all quizzes for student's class
    quizzes = Quiz.objects.filter(
        subject__school_class=student.student_class,
        created_by__role__in=["teacher", "admin", "superadmin"]
    ).select_related("subject", "created_by")

    # Map quiz attempts
    attempts = StudentQuizAttempt.objects.filter(student=student).select_related("quiz")
    attempts_map = {a.quiz_id: a for a in attempts}

    quizzes_with_status = []
    for quiz in quizzes:
        attempt = attempts_map.get(quiz.id)

        if attempt and attempt.is_submitted:
            if quiz.allow_retake:
                status = "Available for Retake"
            else:
                status = "Completed"
        else:
            status = "Not Started"

        quizzes_with_status.append({
            "quiz": quiz,
            "status": status
        })

    return quizzes_with_status



# -----------------------------Retake Approval & Request---------------------------------#


@login_required
@user_passes_test(is_admin_or_superadmin)
def retake_requests_list(request):
    """
    List retake requests; supports AJAX pagination/search returning HTML fragment.
    """
    q = request.GET.get("q", "").strip()
    page_number = request.GET.get("page", 1)

    requests_qs = RetakeRequest.objects.select_related("student", "quiz").order_by("-created_at")
    if q:
        requests_qs = requests_qs.filter(
            Q(student__first_name__icontains=q) |
            Q(student__last_name__icontains=q) |
            Q(quiz__title__icontains=q)
        )

    paginator = Paginator(requests_qs, 10)
    requests_page = paginator.get_page(page_number)

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        html = render_to_string("exams/partials/_retake_requests_table.html", {"requests": requests_page}, request=request)
        return HttpResponse(html)

    return render(request, "exams/admin_retake_requests.html", {"requests": requests_page})


@login_required
@user_passes_test(is_admin_or_superadmin)
@require_POST
def handle_retake_request(request, request_id):
    """
    Approve or deny a retake request (POST only).
    Expects POST param: decision = "approve" | "deny"
    """
    req = get_object_or_404(RetakeRequest.objects.select_related("student", "quiz"), id=request_id)

    # prevent double-processing
    if req.status != "pending":
        return JsonResponse({"success": False, "message": "Request already processed"}, status=400)

    decision = request.POST.get("decision")
    if decision not in ("approve", "deny"):
        return JsonResponse({"success": False, "message": "Invalid decision"}, status=400)

    if decision == "approve":
        # mark request approved
        req.status = "approved"
        req.save(update_fields=["status"])

        # find the previous attempt to remove (prefer explicit relation if present)
        attempt = getattr(req, "attempt", None)
        if not attempt:
            attempt = StudentQuizAttempt.objects.filter(student=req.student, quiz=req.quiz).order_by("-submitted_at").first()

        if attempt:
            # delete answers then attempt
            Answer.objects.filter(attempt=attempt).delete()
            attempt.delete()

        message = f"Your retake request for {req.quiz.title} was approved."
        # log & notify
        ActionLog.objects.create(
            user=request.user,
            action_type="Retake Approved",
            model_name="RetakeRequest",
            object_id=str(req.id),
            details={"student": req.student.username, "exam": req.quiz.title},
        )
        Notification.objects.create(
            sender=request.user,
            recipient=req.student,
            role="student",
            message=message,
        )

        return JsonResponse({"success": True, "message": "Retake approved and previous attempt removed"})

    # deny branch
    req.status = "denied"
    req.save(update_fields=["status"])
    message = f"Your retake request for {req.quiz.title} was denied."
    ActionLog.objects.create(
        user=request.user,
        action_type="Retake Denied",
        model_name="RetakeRequest",
        object_id=str(req.id),
        details={"student": req.student.username, "exam": req.quiz.title},
    )
    Notification.objects.create(
        sender=request.user,
        recipient=req.student,
        role="student",
        message=message,
    )
    return JsonResponse({"success": True, "message": "Retake denied"})

# -----------------------------Retake Approval & Request ended ---------------------------------#


# --------------------------------------------------------------#
# PDF export (consolidated results for a student)

# def download_student_full_report(request, student_id):
#     """Download full report for a single student (all quizzes)."""
#     # Fetch attempts for this student
#     attempts = StudentQuizAttempt.objects.filter(student_id=student_id).select_related("quiz", "quiz__subject")

#     # Prepare PDF response
#     response = HttpResponse(content_type="application/pdf")
#     response["Content-Disposition"] = f'attachment; filename="student_{student_id}_results.pdf"'

#     doc = SimpleDocTemplate(response, pagesize=A4)
#     styles = getSampleStyleSheet()
#     elements = []

#     # Header: logo | school name + student | student photo
#     school_logo = os.path.join(settings.MEDIA_ROOT, "school_logo.png")
#     student = attempts.first().student if attempts.exists() else None
#     school_name = getattr(settings, "SCHOOL_NAME", "My School Name")
#     student_photo = getattr(student, "profile_picture", None)  # assuming User has photo field

#     header_data = [
#         [
#             Image(school_logo, width=50, height=50) if os.path.exists(school_logo) else "",
#             Paragraph(f"<b>{school_name}</b><br/> " + (student.get_full_name() if student else "Unknown"), styles["Title"]),
#             Image(student_photo.path, width=50, height=50) if student_photo and os.path.exists(student_photo.path) else "",
#         ]
#     ]
#     header_table = Table(header_data, colWidths=[70, 350, 70])
#     header_table.setStyle(TableStyle([("ALIGN", (1, 0), (1, 0), "CENTER")]))
#     elements.append(header_table)
#     elements.append(Spacer(1, 20))

#     # Loop through attempts
#     for attempt in attempts:
#         elements.append(Paragraph(f"<b>Exam:</b> {attempt.quiz.title}", styles["Heading3"]))
#         elements.append(Paragraph(f"Subject: {attempt.quiz.subject.name}", styles["Normal"]))
#         elements.append(Paragraph(f"Date: {attempt.started_at.strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))

#         # Collect answers
#         answers = Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice")

#         data = [["Question", "Answer", "Marks", "Feedback"]]
#         for ans in answers:
#             if ans.selected_choice:
#                 ans_text = ans.selected_choice.text
#             else:
#                 ans_text = ans.text_answer or "-"
#             data.append([ans.question.text, ans_text, ans.score, ans.feedback or ""])

#         table = Table(data, colWidths=[200, 150, 60, 100])
#         table.setStyle(TableStyle([
#             ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
#             ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
#         ]))
#         elements.append(table)
#         elements.append(Spacer(1, 15))

#     doc.build(elements)
#     return response



from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from django.http import HttpResponse
from django.conf import settings
from datetime import datetime
import os

def generate_student_report(student, attempts):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # === STYLING HELPERS ===
    def draw_image_with_white_bg(img_path, x, y, w=70, h=70):
        """Draws an image with white rounded background to reduce black border."""
        c.setFillColor(colors.white)
        c.roundRect(x - 5, y - 5, w + 20, h + 20, 10, fill=True, stroke=False)
        try:
            img = ImageReader(img_path)
            c.drawImage(img, x, y, width=w, height=h, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    # # === HEADER ===
    
    logo_path = os.path.join(settings.BASE_DIR, 'static/images/', 'school_logo2.png')
    student_photo = getattr(student, 'profile_picture', None)
    if student_photo and hasattr(student_photo, 'path'):
        student_photo = student_photo.path

    # Left: Logo
    if os.path.exists(logo_path):
        draw_image_with_white_bg(logo_path, 50, height - 110)

    # Center: School Name
    c.setFont("Helvetica-Bold", 18)
    c.setFillColor(colors.darkblue)
    c.drawCentredString(width / 2, height - 60, getattr(settings, 'SCHOOL_NAME', 'My School Name'))

    # Right: Student Photo
    if student_photo and os.path.exists(student_photo):
        draw_image_with_white_bg(student_photo, width - 120, height - 110)

    # === STUDENT INFO ===
    c.setFont("Helvetica", 12)
    c.setFillColor(colors.black)
    y_info = height - 150
    c.drawString(50, y_info, f"Student Name: {student.get_full_name()}")
    c.drawString(50, y_info - 20, f"Username: {student.username}")
    c.drawString(50, y_info - 40, f"Total Attempts: {len(attempts)}")

    y = y_info - 80

    # === HANDLE NO ATTEMPTS ===
    if not attempts.exists():
        c.setFont("Helvetica-Oblique", 13)
        c.setFillColor(colors.red)
        c.drawCentredString(width / 2, y, "No quiz attempts found for this student.")
    else:
        # === TABLE HEADER ===
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(colors.darkblue)
        c.drawString(60, y, "Subject")
        c.drawString(250, y, "Score (%)")
        c.drawString(400, y, "Date Taken")
        c.setFillColor(colors.black)
        c.line(50, y - 5, width - 50, y - 5)
        y -= 25

        # === TABLE ROWS ===
        c.setFont("Helvetica", 11)
        for attempt in attempts:
            if y < 100:
                c.showPage()
                y = height - 100

            subject_name = getattr(attempt.quiz.subject, 'name', str(attempt.quiz.subject))
            c.drawString(60, y, subject_name)
            c.drawString(250, y, f"{attempt.score:.2f}")
            c.drawString(400, y, attempt.submitted_at.strftime("%d-%m-%Y"))
            y -= 20

    # === FOOTER ===
    c.setFont("Helvetica-Oblique", 10)
    c.setFillColor(colors.gray)
    c.drawCentredString(
        width / 2,
        50,
        f"Generated by School Management & CBT App © {datetime.today().year}"
    )

    c.save()
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


def student_report(request, student_id):
    student = request.user.__class__.objects.get(id=student_id)
    attempts = StudentQuizAttempt.objects.filter(student=student)

    pdf = generate_student_report(student, attempts)
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{student.username}_report.pdf"'
    return response


# def generate_student_report(student, attempts):
#     buffer = BytesIO()
#     c = canvas.Canvas(buffer, pagesize=A4)
#     width, height = A4

#     # === HEADER ===
#     logo_path = os.path.join(settings.BASE_DIR, 'static/images/', 'school_logo2.png')
#     student_photo = student.profile_picture.path if hasattr(student, 'profile_picture') and student.profile_picture else None

#     # Left: Logo
#     if os.path.exists(logo_path):
#         draw_image_with_white_bg(logo_path, 50, height - 110)
#         c.drawImage(logo_path, 50, height - 100, width=70, height=70, preserveAspectRatio=True)

#     # Center: School Name
#     c.setFont("Helvetica-Bold", 16)
#     c.drawCentredString(width / 2, height - 60, f"{getattr(settings, 'SCHOOL_NAME', 'My School Name')}")

#     # Right: Student photo
#     if student_photo and os.path.exists(student_photo):
#         c.drawImage(student_photo, width - 120, height - 100, width=70, height=70, preserveAspectRatio=True)

#     # === STUDENT INFO ===
#     c.setFont("Helvetica", 12)
#     c.drawString(50, height - 140, f"Student Name: {student.get_full_name()}")
#     c.drawString(50, height - 160, f"Username: {student.username}")
#     c.drawString(50, height - 180, f"Total Attempts: {len(attempts)}")

#     y = height - 220
#     c.setFont("Helvetica-Bold", 12)
#     c.setFillColor(colors.darkblue)
#     c.drawString(50, y, "Subject")
#     c.drawString(250, y, "Score (%)")
#     c.drawString(400, y, "Date Taken")
#     c.setFillColor(colors.black)
#     c.line(45, y - 5, width - 45, y - 5)
#     y -= 25

#     # === IF NO ATTEMPTS ===
#     if not attempts.exists():
#         c.setFont("Helvetica-Oblique", 12)
#         c.setFillColor(colors.red)
#         c.drawCentredString(width / 2, y - 20, "No exam attempts found yet.")
#     else:
#         # === TABLE ROWS ===
#         c.setFont("Helvetica", 11)
#         for attempt in attempts:
#             if y < 100:  # new page if needed
#                 c.showPage()
#                 y = height - 100

#             subject_name = getattr(attempt.quiz.subject, 'name', str(attempt.quiz.subject))
#             c.drawString(50, y, subject_name)
#             c.drawString(250, y, f"{attempt.score:.2f}")
#             c.drawString(400, y, attempt.submitted_at.strftime("%d-%m-%Y"))
#             y -= 20

#     # === FOOTER ===
#     c.setFont("Helvetica-Oblique", 10)
#     c.setFillColor(colors.black)
#     c.drawCentredString(width / 2, 50, f"Generated by School Management & CBT App © {datetime.today().year}")

#     c.save()
#     pdf = buffer.getvalue()
#     buffer.close()
#     return pdf


# def student_report(request, student_id):
#     student = request.user.__class__.objects.get(id=student_id)
#     attempts = StudentQuizAttempt.objects.filter(student=student)

#     pdf = generate_student_report(student, attempts)
#     response = HttpResponse(pdf, content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename="{student.username}_report.pdf"'
#     return response


def download_closed_quiz_report(request, quiz_id):
    """Download report for all students who attempted a closed quiz."""
    quiz = Quiz.objects.get(id=quiz_id)
    attempts = StudentQuizAttempt.objects.filter(quiz=quiz).select_related("student")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="quiz_{quiz_id}_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Header
    school_logo = os.path.join(settings.MEDIA_ROOT, "school_logo.png")
    school_name = getattr(settings, "SCHOOL_NAME", "My school name")
    header_data = [
        [
            Image(school_logo, width=50, height=50) if os.path.exists(school_logo) else "",
            Paragraph(f"<b>{school_name}</b><br/>Exam Report: {quiz.title}", styles["Title"]),
            "",
        ]
    ]
    header_table = Table(header_data, colWidths=[70, 350, 70])
    header_table.setStyle(TableStyle([("ALIGN", (1, 0), (1, 0), "CENTER")]))
    elements.append(header_table)
    elements.append(Spacer(1, 20))

    # Loop through each student's attempt
    for attempt in attempts:
        student = attempt.student
        elements.append(Paragraph(f"<b>Student:</b> {student.get_full_name()} ({getattr(student, 'student_class', 'N/A')})", styles["Heading3"]))
        elements.append(Paragraph(f"Date: {attempt.started_at.strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))

        answers = Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice")

        data = [["Question", "Answer", "Marks", "Feedback"]]
        for ans in answers:
            if ans.selected_choice:
                ans_text = ans.selected_choice.text
            else:
                ans_text = ans.text_answer or "-"
            data.append([ans.question.text, ans_text, ans.obtained_marks, ans.feedback or ""])

        table = Table(data, colWidths=[200, 150, 60, 100])
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 15))

    doc.build(elements)
    return response


# -------------------------
# Leaderboard (global top students by average score)
# -------------------------
@login_required
def leaderboard(request):
    """
    Renders a leaderboard template showing top students by average score across their completed attempts.
    """
    # compute average attempt scores per student; use StudentQuizAttempt.score (we ensure it's set on submit)
    qs = StudentQuizAttempt.objects.filter(is_submitted=True).values("student").annotate(
        avg_score=Avg("score"), attempts=Sum("score")
    ).order_by("-avg_score")[:50]

    # map to user objects + scores
    leaderboard_list = []
    for item in qs:
        user_id = item["student"]
        try:
            user = User.objects.get(id=user_id)
            leaderboard_list.append({"student": user, "avg_score": item["avg_score"] or 0})
        except User.DoesNotExist:
            continue

    return render(request, "exams/leaderboard.html", {"leaderboard": leaderboard_list})



# ---- Page: create quiz (manual + excel upload) ----

def is_teacher_or_admin(user):
    return user.is_authenticated and user.role in ('teacher','admin','superadmin')


@login_required
@user_passes_test(is_teacher_or_admin)
def create_quiz_page(request):
    subjects = Subject.objects.select_related('school_class').all()
    return render(request, "exams/create_quiz.html", {"subjects": subjects})

# ---- AJAX JSON create ----
@login_required
@user_passes_test(is_teacher_or_admin)
@require_POST
def create_quiz_ajax(request):
    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    title = payload.get("title")
    subject_id = payload.get("subject_id")
    start_time = payload.get("start_time")
    end_time = payload.get("end_time")
    duration_minutes = payload.get("duration_minutes") or 30
    is_published = bool(payload.get("is_published", False))
    questions = payload.get("questions", [])

    if not title or not subject_id:
        return JsonResponse({"ok": False, "error": "Title and subject are required."}, status=400)
    try:
        subject = Subject.objects.get(id=subject_id)
    except Subject.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Subject not found."}, status=404)

    # parse datetimes (accept ISO or 'YYYY-MM-DDTHH:MM' from datetime-local)
    def parse_dt(s):
        if not s:
            return None
        try:
            if "T" in s:
                s = s.replace("T", " ")
            return timezone.make_aware(datetime.strptime(s, "%Y-%m-%d %H:%M"))
        except Exception:
            try:
                return timezone.make_aware(datetime.fromisoformat(s))
            except Exception:
                return None

    st = parse_dt(start_time)
    et = parse_dt(end_time)
    if not st or not et:
        return JsonResponse({"ok": False, "error": "Invalid start_time or end_time format. Use YYYY-MM-DD HH:MM"}, status=400)

    # Build quiz within transaction
    try:
        with transaction.atomic():
            quiz = Quiz.objects.create(
              title=title,
              subject=subject,
              school_class=subject.school_class,  
              created_by=request.user,
              start_time=st,
              end_time=et,
              duration_minutes=int(duration_minutes),
              is_published=is_published
          )
            for qidx, q in enumerate(questions):
                qtext = q.get("text")
                qtype = q.get("question_type")
                qmarks = q.get("marks", 1)
                if not qtext or qtype not in ("objective","subjective"):
                    raise ValueError(f"Invalid question at index {qidx}")

                question = Question.objects.create(
                    quiz=quiz,
                    text=qtext,
                    question_type=qtype,
                    marks=int(qmarks)
                )

                if qtype == "objective":
                    choices = q.get("choices", [])
                    if not choices or not isinstance(choices, list):
                        raise ValueError(f"Objective question requires choices at index {qidx}")
                    correct_present = False
                    for cidx, c in enumerate(choices):
                        ctext = c.get("text")
                        cis = bool(c.get("is_correct", False))
                        if not ctext:
                            raise ValueError(f"Choice text missing for question {qidx} choice {cidx}")
                        Choice.objects.create(question=question, text=ctext, is_correct=cis)
                        if cis:
                            correct_present = True
                    if not correct_present:
                        raise ValueError(f"At least one correct choice required for question {qidx}")
            # success
            return JsonResponse({"ok": True, "quiz_id": quiz.id, "message": "Exam created."})
    except ValueError as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    except Exception as e:
        return JsonResponse({"ok": False, "error": "Server error: " + str(e)}, status=500)

# ---- Excel import ----
@login_required
@user_passes_test(is_teacher_or_admin)
@require_POST
def import_quiz_excel(request):
    """
    Expect FormData with 'excel_file' and optionally 'subject_id' (or subject/class declared in file).
    Excel format (first sheet):
      - Row 1..6: metadata key/value in column A/B:
         A1: quiz_title  B1: My Quiz Title
         A2: class_name  B2: JSS1
         A3: subject_name B3: Mathematics
         A4: start_time  B4: 2025-09-13 14:00
         A5: end_time    B5: 2025-09-13 15:00
         A6: duration_minutes B6: 60
         A7: is_published B7: True
      - Row 9 header then rows from 10:
         Columns: question_text | question_type | marks | choice_1 | choice_1_correct (0/1) | choice_2 | choice_2_correct | ...
    """
    f = request.FILES.get('excel_file')
    if not f:
        return JsonResponse({"ok": False, "error": "No file uploaded"}, status=400)

    # save to temp location (openpyxl can read file-like objects but safer to store)
    try:
        wb = load_workbook(filename=f, data_only=True)
    except Exception as e:
        return JsonResponse({"ok": False, "error": "Invalid Excel file: " + str(e)}, status=400)

    sheet = wb.active

    def cell_value(r,c):
        return sheet.cell(row=r, column=c).value

    # read metadata
    try:
        meta = {}
        for r in range(1, 9):
            key = cell_value(r,1)
            val = cell_value(r,2)
            if key:
                meta[str(key).strip().lower()] = val
    except Exception:
        return JsonResponse({"ok": False, "error": "Failed to read metadata"}, status=400)

    title = meta.get("exam_title") or meta.get("title")
    class_name = meta.get("class_name")
    subject_name = meta.get("subject_name")
    start_time_s = meta.get("start_time")
    end_time_s = meta.get("end_time")
    duration = int(meta.get("duration_minutes") or meta.get("duration") or 30)
    published = bool(meta.get("is_published")) if meta.get("is_published") is not None else False

    # validate subject/class
    if not (class_name and subject_name and title):
        return JsonResponse({"ok": False, "error": "Metadata must include class_name, subject_name and exam_title"}, status=400)

    try:
        school_class = Class.objects.get(name__iexact=str(class_name).strip())
    except Class.DoesNotExist:
        return JsonResponse({"ok": False, "error": f"Class '{class_name}' not found"}, status=404)
    try:
        subject = Subject.objects.get(name__iexact=str(subject_name).strip(), school_class=school_class)
    except Subject.DoesNotExist:
        return JsonResponse({"ok": False, "error": f"Subject '{subject_name}' not found for class {class_name}"}, status=404)

    # parse datetimes (support datetime objects too)
    def parse_dt_obj(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return timezone.make_aware(v) if timezone.is_naive(v) else v
        try:
            s = str(v)
            if "T" in s: s = s.replace("T"," ")
            return timezone.make_aware(datetime.strptime(s, "%d-%m-%Y %H:%M"))
        except Exception:
            try:
                return timezone.make_aware(datetime.fromisoformat(str(v)))
            except Exception:
                return None

    st = parse_dt_obj(start_time_s)
    et = parse_dt_obj(end_time_s)
    if not st or not et:
        return JsonResponse({"ok": False, "error": "Invalid start_time or end_time in metadata. Use 'DD-MM-YYYY HH:MM' or Excel datetime."}, status=400)

    # find header row (we expect header at row 9 or row with 'question_text')
    header_row = None
    for r in range(1, 30):
        first_col = cell_value(r,1)
        if first_col and str(first_col).strip().lower() in ("question_text","question","q_text"):
            header_row = r
            break
    if header_row is None:
        # assume header at row 9
        header_row = 9

    # parse questions starting from header_row+1 to last row with question_text
    questions = []
    r = header_row + 1
    while True:
        q_text = cell_value(r, 1)
        if q_text is None:
            break
        q_type = cell_value(r, 2) or 'objective'
        q_marks = int(cell_value(r, 3) or 1)

        # choices start at col 4, every pair: text, correct
        choices = []
        col = 4
        while True:
            c_text = cell_value(r, col)
            c_flag = cell_value(r, col+1)
            if c_text is None:
                break
            is_corr = False
            if c_flag in (1, '1', True, 'TRUE', 'true'):
                is_corr = True
            choices.append({"text": str(c_text).strip(), "is_correct": bool(is_corr)})
            col += 2

        questions.append({
            "text": str(q_text).strip(),
            "question_type": str(q_type).strip().lower(),
            "marks": q_marks,
            "choices": choices
        })
        r += 1

    # Now create quiz
    try:
        with transaction.atomic():
            quiz = Quiz.objects.create(
                title=title,
                school_class=school_class,
                subject=subject,
                created_by=request.user,
                start_time=st,
                end_time=et,
                duration_minutes=duration,
                is_published=published
            )
            # create questions
            for q in questions:
                qtext = q["text"]
                qtype = q["question_type"]
                qmarks = q["marks"]
                question = Question.objects.create(quiz=quiz, text=qtext, question_type=qtype, marks=qmarks)
                if qtype == "objective":
                    if not q["choices"]:
                        raise ValueError("Objective question without choices found in Excel.")
                    correct_found = False
                    for c in q["choices"]:
                        Choice.objects.create(question=question, text=c["text"], is_correct=c["is_correct"])
                        if c["is_correct"]:
                            correct_found = True
                    if not correct_found:
                        raise ValueError("Objective question must have at least one correct choice.")
            return JsonResponse({"ok": True, "quiz_id": quiz.id, "message": "Exam imported from Excel."})
    except ValueError as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    except Exception as e:
        return JsonResponse({"ok": False, "error": "Import failed: " + str(e)}, status=500)


# ---- Manage quizzes page ----
@login_required
@user_passes_test(is_teacher_or_admin)
def manage_quizzes_page(request):
    if request.user.role == 'teacher':
        quizzes = Quiz.objects.filter(created_by=request.user).order_by('-created_at')
    else:
        quizzes = Quiz.objects.all().order_by('-created_at')
    return render(request, "exams/manage_quizzes.html", {"quizzes": quizzes})

# ---- AJAX publish toggle ---old first----#
@login_required
@user_passes_test(is_teacher_or_admin)
@require_POST
def publish_toggle_ajax(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    # only creator or admin can toggle
    if request.user != quiz.created_by and request.user.role not in ('admin','superadmin'):
        return JsonResponse({"ok": False, "error": "Permission denied."}, status=403)
    quiz.is_published = not quiz.is_published
    quiz.save()
    return JsonResponse({"ok": True, "is_published": quiz.is_published})

# ---- AJAX delete ----
@login_required
@user_passes_test(is_teacher_or_admin)
@require_POST
def delete_quiz_ajax(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    if request.user != quiz.created_by and request.user.role not in ('admin','superadmin'):
        return JsonResponse({"ok": False, "error": "Permission denied."}, status=403)
    quiz.delete()
    ActionLog.objects.create(
    user=request.user,
    action_type= "Delete Exam",
    model_name="Exam",
    object_id=str(quiz.id),
    details={"title": quiz.title, "subject": quiz.subject.name, "action": "deleted"},
)

    return JsonResponse({"ok": True})



def is_teacher_or_admin(user):
    return user.is_authenticated and user.role in ('teacher','admin','superadmin')

@login_required
@user_passes_test(is_teacher_or_admin)
def edit_quiz_page(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    # permission: only creator or admin allowed to edit
    if request.user != quiz.created_by and request.user.role not in ('admin','superadmin'):
        return HttpResponse("Permission denied", status=403)
    subjects = Subject.objects.select_related('school_class').all()

    # We'll render a page similar to create_quiz but include a small script that fetches quiz JSON to prefill
    return render(request, "exams/edit_quiz.html", {"quiz": quiz, "subjects": subjects})



@login_required
@user_passes_test(is_teacher_or_admin)
@require_POST
def edit_quiz_ajax(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    if request.user != quiz.created_by and request.user.role not in ('admin','superadmin'):
        return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)

    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    title = payload.get("title")
    subject_id = payload.get("subject_id")
    start_time = payload.get("start_time")
    end_time = payload.get("end_time")
    duration_minutes = payload.get("duration_minutes") or 30
    is_published = bool(payload.get("is_published", False))
    questions = payload.get("questions", [])

    if not title or not subject_id:
        return JsonResponse({"ok": False, "error": "Title and subject are required."}, status=400)

    try:
        subject = Subject.objects.get(id=subject_id)
    except Subject.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Subject not found."}, status=404)

    # Update within transaction. We'll replace existing questions/choices with new ones.
    try:
        with transaction.atomic():
            quiz.title = title
            quiz.subject = subject
            # parse start/end strings same as in create; expecting ISO or "YYYY-MM-DD HH:MM"
            from django.utils.dateparse import parse_datetime

            start_time = parse_datetime(payload.get("start_time"))
            end_time = parse_datetime(payload.get("end_time"))

            if not start_time or not end_time:
                return JsonResponse({"ok": False, "error": "Invalid start/end time format."}, status=400) 
            
            quiz.start_time = start_time
            quiz.end_time = end_time
            quiz.duration_minutes = int(duration_minutes)
            quiz.is_published = is_published
            quiz.save()
            

            # delete old questions & choices, then recreate
            quiz.questions.all().delete()
            for qidx, q in enumerate(questions):
                qtext = q.get("text")
                qtype = q.get("question_type")
                qmarks = q.get("marks", 1)
                if not qtext or qtype not in ("objective","subjective"):
                    raise ValueError(f"Invalid question at index {qidx}")
                question = Question.objects.create(
                    quiz=quiz, text=qtext, question_type=qtype, marks=int(qmarks)
                )
                if qtype == "objective":
                    choices = q.get("choices", [])
                    if not choices or not isinstance(choices, list):
                        raise ValueError(f"Objective question requires choices at index {qidx}")
                    correct_found = False
                    for cidx, c in enumerate(choices):
                        ctext = c.get("text")
                        cis = bool(c.get("is_correct", False))
                        if not ctext:
                            raise ValueError(f"Choice text missing for question {qidx} choice {cidx}")
                        Choice.objects.create(question=question, text=ctext, is_correct=cis)
                        if cis:
                            correct_found = True
                    if not correct_found:
                        raise ValueError(f"At least one correct choice required for question {qidx}")
            # success
            ActionLog.objects.create(
            user=request.user,
            action_type="Edit Exam",
            model_name="Exam",
            object_id=str(quiz.id),
            details={"title": quiz.title, "subject": quiz.subject.name, "action": "edited" },
        )

        return JsonResponse({"ok": True, "quiz_id": quiz.id, "message": "Exam updated."})
    
    except ValueError as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)
    except Exception as e:
        return JsonResponse({"ok": False, "error": "Server error: " + str(e)}, status=500)



@require_POST
def toggle_quiz_publish(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    quiz.is_published = not quiz.is_published
    quiz.save()
    return JsonResponse({"ok": True, "new_status": quiz.is_published})

def quiz_detail_api(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    return JsonResponse({
        "ok": True,
        "quiz": {
            "id": quiz.id,
            "title": quiz.title,
            "subject": str(quiz.subject),
            "class_name": str(quiz.subject.school_class),
            "duration_minutes": quiz.duration_minutes,
            "is_published": quiz.is_published,
        }
    })


@login_required
@user_passes_test(is_teacher_or_admin)
def quiz_details_page(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    # permission: only creator or admin allowed to view details
    if request.user != quiz.created_by and request.user.role not in ('admin','superadmin'):
        return HttpResponse("Permission denied", status=403)
    if quiz.DoesNotExist:
        attempts = StudentQuizAttempt.objects.filter(quiz=quiz).select_related('student').order_by('-started_at')
        return redirect('exams:quiz_closed_detail', quiz.id)
    
    attempts = StudentQuizAttempt.objects.filter(quiz=quiz).select_related('student').order_by('-started_at')

    return render(request, "exams/quiz_details.html", {"quiz": quiz, "attempts": attempts})


def manage_quizzes_redirect(request):
    if request.user.role == 'admin' or 'superadmin':
        return redirect('exams:manage_quizzes_admin')
    elif request.user.role == 'teacher':
        return redirect('exams:manage_quizzes_teacher')
    else:
        messages.warning(request, 'Unknown user role')
        return redirect('users:login')



def quiz_closed(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    
    # Make sure quiz is actually closed
    if quiz.end_time > timezone.now():
        # If not yet closed, you can redirect to quiz detail or show error
        return render(request, "exams/not_closed.html", {"quiz": quiz})
    
    # Get all attempts
    attempts = StudentQuizAttempt.objects.filter(quiz=quiz).select_related("student").order_by("-started_at")
    context = {
        "quiz": quiz,
        "attempts": attempts,
    }
    return render(request, "exams/quiz_closed.html", context)




from openpyxl import Workbook
from django.http import HttpResponse

@login_required
@user_passes_test(is_teacher_or_admin)
def download_excel_template(request):
    now = datetime.now()
    wb = Workbook()
    ws = wb.active
    ws.title = "ExamTemplate"

    # metadata rows
    ws['A1'] = 'exam_title'; ws['B1'] = "FIRST TERM EXAM 2025/2026 SESSION" 
    ws['A2'] = 'class_name'; ws['B2'] = 'JSS1'
    ws['A3'] = 'subject_name'; ws['B3'] = 'Mathematics'
    ws['A4'] = 'start_time'; ws['B4'] = now
    ws['A5'] = 'end_time'; ws['B5'] = now
    ws['A6'] = 'duration_minutes'; ws['B6'] = 60
    ws['A7'] = 'is_published'; ws['B7'] = 'True'

    # header row at row 9
    headers = ['question_text','question_type','marks']
    # allow up to 6 choices (choice_x, choice_x_correct)
    for i in range(1,7):
        headers.append(f'choice_{i}')
        headers.append(f'choice_{i}_correct')
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=9, column=col_idx, value=header)

    # example question row
    ws.cell(row=10, column=1, value='What is 2+2?')
    ws.cell(row=10, column=2, value='objective')
    ws.cell(row=10, column=3, value='1')
    ws.cell(row=10, column=4, value='3')
    ws.cell(row=10, column=5, value='0')
    ws.cell(row=10, column=6, value='4')
    ws.cell(row=10, column=7, value='1')

    # prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Exam_template.xlsx'
    wb.save(response)
    return response


@login_required
def quiz_json3(request, attempt_id):
    """
    Return quiz questions + answers as JSON for a student's attempt.
    Supports resume if attempt not submitted.
    """

    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id, student=request.user)

    # Check expiry
    if attempt.end_time and timezone.now() > attempt.end_time:
        return JsonResponse({"error": "Exam time expired."}, status=403)

    # Check submission
    if attempt.is_submitted:
        return JsonResponse({"error": "Exam already submitted."}, status=403)
    quiz = attempt.quiz
    questions = quiz.questions.prefetch_related("choice_set")

    data = []
    for q in questions:
        saved_answer = Answer.objects.filter(attempt=attempt, question=q).first()
        data.append({
            "id": q.id,
            "text": q.text,
            "type": q.question_type,
            "marks": q.marks,
            "choices": [{"id": c.id, "text": c.text} for c in q.choice_set.all()],
            "saved_choice": saved_answer.selected_choice.id if saved_answer and saved_answer.selected_choice else None,
            "saved_text": saved_answer.text_answer if saved_answer else "",
        })


    return JsonResponse({
        "quiz": {
            "id": quiz.id,
            "title": quiz.title,
            "subject": quiz.subject.name,
            "duration_minutes": quiz.duration,
            "end_time": attempt.end_time.isoformat() if attempt.end_time else None,
        },
        "questions": data,
    })




@login_required
def quiz_json1(request, quiz_id):
    quiz = get_object_or_404(Quiz.objects.prefetch_related("choices"), id=quiz_id)

    data = {
        "quiz": {
            "id": quiz.id,
            "title": quiz.title,
            "end_time": quiz.end_time.isoformat() if quiz.end_time else None,
            "questions": []
        }
    }

    for q in quiz.questions.all():
        data["quiz"]["questions"].append({
            "id": q.id,
            "text": q.text,
            "question_type": q.question_type,
            "marks": q.marks,
            "choices": [
                {"id": c.id, "text": c.text} for c in q.choices.all()
            ]
        })

    return JsonResponse(data)


@login_required

def quiz_json_quiz_load(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)

    # permission: only creator or admin/superadmin can fetch edit JSON
    if request.user != quiz.created_by and request.user.role not in ('admin','superadmin'):
        return JsonResponse({"ok": False, "error": "Permission denied"}, status=403)

    quiz_data = {
        "id": quiz.id,
        "title": quiz.title,
        "subject_id": quiz.subject.id if quiz.subject else None,
        "subject_name": quiz.subject.name if quiz.subject else None,
        "start_time": quiz.start_time.strftime("%Y-%m-%d %H:%M") if quiz.start_time else None,
        "end_time": quiz.end_time.strftime("%Y-%m-%d %H:%M") if quiz.end_time else None,
        "duration_minutes": quiz.duration_minutes,
        "is_published": quiz.is_published,
        "questions": []
    }
    for q in quiz.questions.all():
        qdata = {
            "id": q.id,
            "text": q.text,
            "question_type": q.question_type,
            "marks": q.marks,
            "choices": []  
        }
        if q.question_type == "objective":
            qdata["choices"] = [
                {"id": c.id, "text": c.text, "is_correct": c.is_correct}
                for c in q.choices.all()
            ]
        quiz_data["questions"].append(qdata)
      

    return JsonResponse({"ok": True, "quiz": quiz_data})


###------- Retake Exam after submission by student this will check their attempt and retake count-------###

@login_required
def request_retake(request, attempt_id):
    """
    Student requests a retake for a submitted attempt.
    """
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id, student=request.user)

    if not attempt.completed:
        messages.error(request, "You can only request a retake after submitting the Exam.")
        return redirect("student_dashboard")

    if getattr(attempt, "retake_requested", False):
        messages.info(request, "You have already requested a retake for this Exam.")
        return redirect("student_dashboard")

    # Mark request
    attempt.retake_requested = True
    attempt.save()

    # Notify admin/teacher
    Notification.objects.create(
        sender=request.user,
        recipient=attempt.quiz.created_by,
        message=f"📩 {request.user.username} has requested a retake for '{attempt.quiz.title}'.",
        role="admin",
        is_broadcast=False,
    )

    messages.success(request, "Your retake request has been sent for approval.")
    return redirect("users:student_dashboard")

###---------------------The End Retake Exam after submission -----------------------###



###------- Retake Exam after submission by student this will check their attempt and retake count-------###

@user_passes_test(is_admin_or_superadmin)
def retake_requests(request): # Retake request button on dashboard
    """
    Admin page: list attempts that are submitted/expired.
    """
    attempts = StudentQuizAttempt.objects.filter(is_submitted=True).select_related("quiz", "student")
    return render(request, "exams/retake_requests.html", {"attempts": attempts})


@user_passes_test(is_admin_or_superadmin)
def approve_retake(request, attempt_id):
    """
    Admin approves retake request.
    """
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id)

    # Reset attempt for retake
    attempt.retake_allowed = True
    attempt.is_submitted = False
    attempt.end_time = None
    attempt.save()

    # Notify studenl
    Notification.objects.create(
        recipient=attempt.student,
        message=f"✅ Your retake request for Exam '{attempt.quiz.title}' has been approved.",
        is_read=False
    )

    # Log action
    ActionLog.objects.create(
        user=request.user,
        action_type="approve_retake",
        details={"attempt_id": attempt.id, "student": attempt.student.username, "Exam": attempt.quiz.title}
    )

    messages.success(request, f"Retake approved for {attempt.student.username} on {attempt.quiz.title}.")
    return redirect("exams:retake_requests")



##------------New Student Exam take/Retake/ etc Page ------------------###



# helper: check student
def is_student(user):
    return user.is_authenticated and getattr(user, "role", None) == "student"

def is_teacher_or_admin(user):
    return user.is_authenticated and getattr(user, "role", None) in ("teacher", "admin", "superadmin")



# ---- Page: take quiz ---- without randomization and with retake allowed ----##
@login_required
def take_quiz_view(request, quiz_id):
    """
    Show the take-quiz page. Create or resume attempt server-side.
    """
    if not is_student(request.user):
        return HttpResponseForbidden("forbidden") 

    quiz = get_object_or_404(Quiz, id=quiz_id)

    now = timezone.now()
    # allowed to start/resume if published OR retake allowed by admin/teacher, etc.
    # find existing active attempt
    attempt = StudentQuizAttempt.objects.filter(student=request.user, quiz=quiz, is_submitted=False).order_by("-started_at").first()

    if attempt:
        # if expired, do not resume (unless retake_allowed)
        if attempt.end_time and now > attempt.end_time and not attempt.retake_allowed:
            attempt = None

    if not attempt:
        # not resuming: check whether quiz open or retake allowed
        available = quiz.is_published and (quiz.start_time <= now <= quiz.end_time)
        last_attempt = StudentQuizAttempt.objects.filter(student=request.user, quiz=quiz).order_by("-started_at").first()
        allow_ret = quiz.allow_retake or (last_attempt and last_attempt.retake_allowed)
        if not available and not allow_ret:
            # if closed and not allowed: if there is a is_submitted attempt, redirect to its result
            if last_attempt and last_attempt.is_submitted:
                return redirect('exams:quiz_result', attempt_id=last_attempt.id)
            return redirect('exams:quiz_closed_detail', quiz_id=quiz.id)

        # create new attempt
        end_time = now + timezone.timedelta(minutes=quiz.duration_minutes)
        attempt = StudentQuizAttempt.objects.create(
            student=request.user,
            quiz=quiz,
            started_at=now,
            end_time=end_time,
            is_submitted=False,
            retake_allowed=False,
            retake_count=(last_attempt.retake_count+1 if last_attempt else 0),
            score=0.0
        )
        ActionLog.objects.create(user=request.user, action_type="Started Exam", description=f"Started exam {quiz.title}", model_name="StudentQuizAttempt", object_id=str(attempt.id), details={"Exam": quiz.title})

    # ensure questions are prefetched for template rendering
    questions = quiz.questions.prefetch_related("choices").all()
    # also pass existing saved answers to prefill (dict question_id -> answer)
    saved_answers = {}
    for ans in Answer.objects.filter(attempt=attempt).select_related("selected_choice", "question"):
        if ans.question.question_type == "objective" and ans.selected_choice:
            saved_answers[str(ans.question.id)] = {"choice_id": ans.selected_choice.id}
        else:
            saved_answers[str(ans.question.id)] = {"text": ans.text_answer}

    context = {
        "quiz": quiz,
        "attempt": attempt,
        "questions": questions,
        "saved_answers": saved_answers,
    }
    return render(request, "exams/take_quiz.html", context)


## ---- Page: take quiz with random question and choice order ---- ##
# import random, json
# from django.shortcuts import render, get_object_or_404, redirect
# from django.utils import timezone
# from django.contrib.auth.decorators import login_required
# from django.http import JsonResponse
# from .models import Quiz, Question, Choice, StudentQuizAttempt, Answer


# @login_required
# def take_quiz_view(request, quiz_id):
#     quiz = get_object_or_404(Quiz, id=quiz_id)
#     student = request.user

#     attempt, created = StudentQuizAttempt.objects.get_or_create(
#         quiz=quiz,
#         student=student,
#         defaults={'start_time': timezone.now()}
#     )

#     # === Handle AJAX autosave ===
#     if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
#         data = json.loads(request.body.decode('utf-8'))
#         answers = data.get('answers', {})
#         for qid, ans_text in answers.items():
#             question = Question.objects.get(id=int(qid))
#             Answer.objects.update_or_create(
#                 attempt=attempt,
#                 question=question,
#                 defaults={'answer_text': ans_text}
#             )
#         return JsonResponse({'status': 'saved', 'timestamp': timezone.now().strftime("%H:%M:%S")})

#     # === Handle full submission ===
#     if request.method == "POST":
#         total_score = 0
#         for ans in Answer.objects.filter(attempt=attempt):
#             if ans.question.question_type == 'objective':
#                 correct = ans.question.choices.filter(is_correct=True).first()
#                 if correct and str(correct.id) == ans.answer_text:
#                     total_score += 1
#         attempt.total_score = total_score
#         attempt.is_submitted = True
#         attempt.end_time = timezone.now()
#         attempt.save()
#         return redirect("student_dashboard")

#     # === Generate random question order ===
#     if created or not attempt.question_order:
#         questions = list(quiz.questions.all())
#         random.shuffle(questions)
#         attempt.question_order = ",".join(str(q.id) for q in questions)
#         attempt.save()

#     # === Load questions in saved order ===
#     question_ids = [int(qid) for qid in attempt.question_order.split(",")]
#     questions = list(Question.objects.filter(id__in=question_ids))
#     questions.sort(key=lambda q: question_ids.index(q.id))

#     # === Randomize choices (without direct assignment) ===
#     for q in questions:
#         q.randomized_choices = list(q.choices.all())  # ✅ new temporary field
#         random.shuffle(q.randomized_choices)          # shuffle just for display

#     total_seconds = quiz.duration_minutes * 60 if hasattr(quiz, "duration_minutes") else 600

#     return render(request, "exams/take_quiz_random_all.html", {
#         "quiz": quiz,
#         "questions": questions,
#         "attempt": attempt,
#         "total_seconds": total_seconds,
#     })




# ### ---- Page: take quiz with random question order only ----##
# import random, json
# from django.shortcuts import render, get_object_or_404, redirect
# from django.utils import timezone
# from django.contrib.auth.decorators import login_required
# from django.http import JsonResponse
# from .models import Quiz, Question, Choice, StudentQuizAttempt, Answer

# @login_required
# def take_quiz_random_questions(request, quiz_id):
#     quiz = get_object_or_404(Quiz, id=quiz_id)
#     student = request.user

#     # Create or get existing attempt
#     attempt, created = StudentQuizAttempt.objects.get_or_create(
#         quiz=quiz,
#         student=student,
#         defaults={'start_time': timezone.now()}
#     )

#     # Handle autosave via AJAX
#     if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
#         data = json.loads(request.body.decode('utf-8'))
#         answers = data.get('answers', {})
#         for qid, ans_text in answers.items():
#             question = Question.objects.get(id=int(qid))
#             Answer.objects.update_or_create(
#                 attempt=attempt,
#                 question=question,
#                 defaults={'answer_text': ans_text}
#             )
#         return JsonResponse({'status': 'saved', 'timestamp': timezone.now().strftime("%H:%M:%S")})

#     # Handle full submission
#     if request.method == "POST":
#         total_score = 0
#         for ans in Answer.objects.filter(attempt=attempt):
#             if ans.question.question_type == 'objective':
#                 correct_choice = ans.question.choice_set.filter(is_correct=True).first()
#                 if correct_choice and str(correct_choice.id) == ans.answer_text:
#                     total_score += 1
#         attempt.total_score = total_score
#         attempt.is_submitted = True
#         attempt.end_time = timezone.now()
#         attempt.save()
#         return redirect("student_dashboard")

#     # Randomize question order (once per attempt)
#     if created or not attempt.question_order:
#         questions = list(quiz.question_set.all())
#         random.shuffle(questions)
#         attempt.question_order = ",".join(str(q.id) for q in questions)
#         attempt.save()

#     # Preserve saved order
#     question_ids = [int(qid) for qid in attempt.question_order.split(",")]
#     questions = list(Question.objects.filter(id__in=question_ids))
#     questions.sort(key=lambda q: question_ids.index(q.id))

#     # Duration in seconds (default 10 mins if not set)
#     total_seconds = quiz.duration * 60 if hasattr(quiz, "duration") else 600

#     return render(request, "exams/take_quiz_random_questions.html", {
#         "quiz": quiz,
#         "questions": questions,
#         "attempt": attempt,
#         "total_seconds": total_seconds,
#     })



@login_required
@require_POST
def api_submit_answer(request, attempt_id):
    """
    Autosave single answer (AJAX). Accepts JSON: {question_id, answer}
    - For objective: answer is choice id (int)
    - For subjective: answer is text
    Autograde objectives here.
    """
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id, student=request.user)
    if attempt.is_submitted:
        return JsonResponse({"ok": False, "error": "Attempt already submitted"}, status=400)
    if attempt.end_time and timezone.now() > attempt.end_time:
        return JsonResponse({"ok": False, "error": "Attempt time expired"}, status=400)

    try:
        payload = json.loads(request.body.decode())
    except Exception:
        return JsonResponse({"ok": False, "error": "invalid json"}, status=400)

    qid = payload.get("question_id")
    answer = payload.get("answer", None)  # can be choice id or text

    if not qid:
        return JsonResponse({"ok": False, "error": "question_id required"}, status=400)

    question = get_object_or_404(Question, id=qid, quiz=attempt.quiz)

    # objective
    if question.question_type == "objective":
        # answer should be a choice id (int or str)
        try:
            choice = Choice.objects.get(id=int(answer), question=question)
        except Exception:
            choice = None

        if choice:
            obtained = float(question.marks) if choice.is_correct else 0.0
            ans_obj, created = Answer.objects.update_or_create(
                attempt=attempt, question=question,
                defaults={
                    "selected_choice": choice,
                    "text_answer": None,
                    "score": obtained,
                    "is_pending": False,
                }
            )
        else:
            # clear selection
            ans_obj, created = Answer.objects.update_or_create(
                attempt=attempt, question=question,
                defaults={
                    "selected_choice": None,
                    "text_answer": None,
                    "score": 0.0,
                    "is_pending": False,
                }
            )
    else:
        # subjective: save text, mark is_pending True, obtained_marks left as 0 (teacher will grade later)
        text = str(answer or "")
        ans_obj, created = Answer.objects.update_or_create(
            attempt=attempt, question=question,
            defaults={
                "selected_choice": None,
                "text_answer": text,
                "score": 0.0,
                "is_pending": True,
            }
        )

    ActionLog.objects.create(user=request.user, action_type="submit_answer", description="Autosaved answer", model_name="Answer", object_id=str(ans_obj.id), details={"question": question.id})
    return JsonResponse({"ok": True})


@login_required
@require_POST
def api_submit_attempt(request, attempt_id):
    """
    Final submission: autograde objective parts (again to be safe), sum objective + graded subjective.
    Returns JSON with overall score and per-question details.
    """
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id, student=request.user)

    if attempt.is_submitted:
        return JsonResponse({"ok": False, "error": "Already submitted"}, status=400)

    # re-evaluate objective answers and ensure answers exist for all objective questions
    answers_qs = Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice")
    # Ensure every objective question has an Answer row (create with 0 if missing)
    for q in attempt.quiz.questions.filter(question_type="objective"):
        ans, created = Answer.objects.get_or_create(attempt=attempt, question=q, defaults={
            "selected_choice": None, "text_answer": None, "score": 0.0, "is_pending": False
        })

    # Auto-grade objective answers
    score = 0.0

    for ans in Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice"):
        q = ans.question
        if q.question_type == "objective":
            if ans.selected_choice and getattr(ans.selected_choice, "is_correct", False):
                ans.score = float(q.marks)
                score += q.marks
            
            else:
                ans.score = 0.0
            ans.is_pending = False
            ans.save(update_fields=["score", "is_pending"])
        # subjective answers left as is (pending until graded)

       # calculate totals using Answer helper methods
    objective_sum = Answer.objective_score(attempt)
    subjective_sum = Answer.subjective_score(attempt)  # only graded subjective answers contribute
    total_score = float(objective_sum) + float(subjective_sum)

    attempt.score = total_score
    attempt.is_submitted = True
    attempt.submitted_at = timezone.now() if hasattr(attempt, "submitted_at") else timezone.now()
    attempt.save()
 

    ActionLog.objects.create(user=request.user, action_type="Submitted Exam", description=f"Submitted attempt {attempt.id}", model_name="StudentQuizAttempt", object_id=str(attempt.id), details={"score": attempt.score})

    # Build a summary payload
    question_details = []
    for ans in Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice"):
        q = ans.question
        if q.question_type == "objective":
            student_ans = ans.selected_choice.text if ans.selected_choice else None
            correct = bool(ans.selected_choice and getattr(ans.selected_choice, "is_correct", False))
            correct_ans_text = ", ".join([c.text for c in q.choices.filter(is_correct=True)])
            question_details.append({
                "question_id": q.id,
                "type": "objective",
                "text": q.text,
                "student": student_ans,
                "correct_answer": correct_ans_text,
                "marks_awarded": ans.score,
                "max_marks": q.marks,
                "status": "correct" if correct else "wrong",
            })
        else:
            question_details.append({
                "question_id": q.id,
                "type": "subjective",
                "text": q.text,
                "student": ans.text_answer,
                "marks_awarded": ans.score if not ans.is_pending else None,
                "max_marks": q.marks,
                "status": "pending" if ans.is_pending else "graded",
                "feedback": ans.feedback,
            })

    return JsonResponse({
        "ok": True,
        "score": total_score,
        "objective_sum": float(objective_sum),
        "subjective_sum": float(subjective_sum),
        "questions": question_details,
    })


@login_required
def quiz_result_view(request, attempt_id):
    """
    Render result page with per-question breakdown and Chart.js bar for objective correctness.
    """
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id)
    # allow only owner or teacher/admin
    if request.user != attempt.student and not is_teacher_or_admin(request.user):
        return HttpResponseForbidden("forbidden")

    answers = Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice", "graded_by")
    qrows = []
    correct_count = 0
    for ans in answers:
        q = ans.question
        if q.question_type == "objective":
            correct = bool(ans.selected_choice and getattr(ans.selected_choice, "is_correct", False))
            if correct:
                correct_count += 1
            correct_text = ", ".join([c.text for c in q.choices.filter(is_correct=True)])
            student_text = ans.selected_choice.text if ans.selected_choice else (ans.text_answer or "")
            qrows.append({
                "id": q.id,
                "text": q.text,
                "type": "objective",
                "student_answer": student_text,
                "correct_answer": correct_text,
                "marks_awarded": ans.score,
                "max_marks": q.marks,
                "status": "correct" if correct else "wrong",
            })
        else:
            qrows.append({
                "id": q.id,
                "text": q.text,
                "type": "subjective",
                "student_answer": ans.text_answer,
                "marks_awarded": (ans.score if not ans.is_pending else None),
                "max_marks": q.marks,
                "status": "pending" if ans.is_pending else "graded",
                "feedback": ans.feedback,
            })

    # chart data for objective questions: percent correct per question (0 or 1)
    chart_labels = []
    chart_values = []
    for row in qrows:
        chart_labels.append(f"Q{row['id']}")
        if row['type'] == 'objective':
            chart_values.append((row['marks_awarded'] or 0) / (row['max_marks'] or 1) * 100)
        else:
            chart_values.append(None)  # will show as gap

    context = {
        "attempt": attempt,
        "qrows": qrows,
        "chart_labels": chart_labels,
        "chart_values": chart_values,
        "total_score": attempt.score,
        "correct_count": correct_count,
        "total_questions": len(qrows),
    }
    return render(request, "exams/quiz_result.html", context)


@login_required
def review_attempt_view(request, attempt_id):
    """
    Review answers + teacher feedback. Owner or teacher/admin only.
    """
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id)
    if request.user != attempt.student and not is_teacher_or_admin(request.user):
        return HttpResponseForbidden("forbidden")

    answers = Answer.objects.filter(attempt=attempt).select_related("question", "selected_choice", "graded_by")
    return render(request, "exams/quiz_review.html", {"attempt": attempt, "answers": answers})



@login_required
@require_POST
def student_request_retake_view(request, quiz_id):
    """
    Student requests retake: creates RetakeRequest and notifies teacher + admins.
    """
    if not is_student(request.user):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    quiz = get_object_or_404(Quiz, id=quiz_id)
  

    # Get reason (from JSON body or POST)
    reason = ""
    try:
        payload = json.loads(request.body.decode())
        reason = payload.get("reason", "")[:1000]
    except Exception:
        reason = request.POST.get("reason", "")[:1000]

    # Create RetakeRequest
    rr = RetakeRequest.objects.create(
        student=request.user,
        quiz=quiz,
        reason=reason,
        status="pending"
    )

    # Notify teacher
    Notification.objects.create(
        sender=request.user,
        recipient=quiz.created_by,
        message=f"Retake request by {request.user.get_full_name()} for {quiz.title}: {reason}",
        role="teacher",
        is_broadcast=False,
    )

    # Notify admins
    admins = User.objects.filter(role__in=("admin", "superadmin"), approved=True)
    for a in admins:
        Notification.objects.create(
            sender=request.user,
            recipient=a,
            message=f"Retake request by {request.user.get_full_name()} for {quiz.title} {quiz.school_class}",
            role="admin",
            is_broadcast=False,
        )

    # Log action
    ActionLog.objects.create(
        user=request.user,
        action_type="retake_request",
        description="Requested retake",
        model_name="RetakeRequest",
        object_id=str(rr.id),
        details={"Exam": quiz.title},
    )

    return JsonResponse({"ok": True, "message": "Retake request sent."})

